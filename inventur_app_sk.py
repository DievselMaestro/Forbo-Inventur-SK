#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
INVENTORY PROGRAM FOR FORBO SK - MALACKY WAREHOUSE MANAGEMENT
==============================================================

Desktop application for warehouse inventory with barcode scanner integration.
Supports Rolls only (no granulate).
Supports two warehouse modes: SK (Malacky) and Zert.
Developed for Windows 11, Python 3.11+

Date: March 2026
Version: 1.1 SK+Zert
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys
import shutil
from pathlib import Path
import json
import logging


# ---------------------------------------------------------------------------
# Warehouse Selection Dialog
# ---------------------------------------------------------------------------

class WarehouseSelectionDialog:
    """Startup dialog: select which warehouse to scan."""

    def __init__(self, parent):
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Lager auswählen / Select Warehouse")
        self.dialog.geometry("620x220")
        self.dialog.resizable(False, False)
        self.dialog.grab_set()
        # Center on screen
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - 310
        y = (self.dialog.winfo_screenheight() // 2) - 110
        self.dialog.geometry(f"+{x}+{y}")
        self.dialog.protocol("WM_DELETE_WINDOW", self._cancel)

        self._build_widgets()
        self.dialog.wait_window()

    def _build_widgets(self):
        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            frame,
            text="Welches Lager scannen?\nWhich warehouse to scan?",
            font=("Arial", 13, "bold"),
            justify=tk.CENTER,
        ).pack(pady=(0, 20))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack()

        tk.Button(
            btn_frame,
            text="HALB",
            font=("Arial", 14, "bold"),
            bg="#1f4e79",
            fg="white",
            width=14,
            height=2,
            command=lambda: self._select("SK"),
        ).pack(side=tk.LEFT, padx=(0, 20))

        tk.Button(
            btn_frame,
            text="ZERT",
            font=("Arial", 14, "bold"),
            bg="#375623",
            fg="white",
            width=14,
            height=2,
            command=lambda: self._select("Zert"),
        ).pack(side=tk.LEFT, padx=(0, 20))

        tk.Button(
            btn_frame,
            text="KMAT",
            font=("Arial", 14, "bold"),
            bg="#6b1f1f",
            fg="white",
            width=14,
            height=2,
            command=lambda: self._select("KMAT"),
        ).pack(side=tk.LEFT, padx=(20, 0))

    def _select(self, mode):
        self.result = mode
        self.dialog.destroy()

    def _cancel(self):
        self.result = None
        self.dialog.destroy()


# ---------------------------------------------------------------------------
# Settings Dialog
# ---------------------------------------------------------------------------

class SettingsDialog:
    """Dialog for configuring file paths (master table + export folder)."""

    def __init__(self, parent, current_config):
        self.result = None
        self.config = current_config.copy()

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Settings")
        self.dialog.geometry("620x580")
        self.dialog.resizable(True, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.geometry(
            "+%d+%d" % (parent.winfo_rootx() + 60, parent.winfo_rooty() + 60)
        )

        self._build_widgets()
        self.dialog.wait_window()

    def _build_widgets(self):
        frame = ttk.Frame(self.dialog, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)

        # --- SK Master Table path ---
        ttk.Label(frame, text="SK Master Table file (*.xlsx):", font=("Arial", 10, "bold")).grid(
            row=0, column=0, sticky=tk.W, pady=(0, 4))

        self.master_path_var = tk.StringVar(
            value=self.config.get("arbeitstabelle_path", ""))
        master_entry = ttk.Entry(frame, textvariable=self.master_path_var, width=55)
        master_entry.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(0, 8))

        ttk.Button(frame, text="Browse...", command=self._browse_master).grid(
            row=1, column=1, sticky=tk.W)

        # --- SK Export folder ---
        ttk.Label(frame, text="SK Export output folder:", font=("Arial", 10, "bold")).grid(
            row=2, column=0, sticky=tk.W, pady=(16, 4))

        self.export_path_var = tk.StringVar(
            value=self.config.get("export_path", ""))
        export_entry = ttk.Entry(frame, textvariable=self.export_path_var, width=55)
        export_entry.grid(row=3, column=0, sticky=(tk.W, tk.E), padx=(0, 8))

        ttk.Button(frame, text="Browse...", command=self._browse_export).grid(
            row=3, column=1, sticky=tk.W)

        # --- Zert Master Table path ---
        ttk.Label(frame, text="Zert Master Table (*.xlsx):", font=("Arial", 10, "bold")).grid(
            row=4, column=0, sticky=tk.W, pady=(16, 4))

        self.zert_master_path_var = tk.StringVar(
            value=self.config.get("arbeitstabelle_zert_path", ""))
        zert_master_entry = ttk.Entry(frame, textvariable=self.zert_master_path_var, width=55)
        zert_master_entry.grid(row=5, column=0, sticky=(tk.W, tk.E), padx=(0, 8))

        ttk.Button(frame, text="Browse...", command=self._browse_zert_master).grid(
            row=5, column=1, sticky=tk.W)

        # --- Zert Export folder ---
        ttk.Label(frame, text="Zert Export folder:", font=("Arial", 10, "bold")).grid(
            row=6, column=0, sticky=tk.W, pady=(16, 4))

        self.zert_export_path_var = tk.StringVar(
            value=self.config.get("export_zert_path", ""))
        zert_export_entry = ttk.Entry(frame, textvariable=self.zert_export_path_var, width=55)
        zert_export_entry.grid(row=7, column=0, sticky=(tk.W, tk.E), padx=(0, 8))

        ttk.Button(frame, text="Browse...", command=self._browse_zert_export).grid(
            row=7, column=1, sticky=tk.W)

        # --- KMAT Master Table path ---
        ttk.Label(frame, text="KMAT Master Table (*.xlsx):", font=("Arial", 10, "bold")).grid(
            row=8, column=0, sticky=tk.W, pady=(16, 4))

        self.kmat_master_path_var = tk.StringVar(
            value=self.config.get("arbeitstabelle_kmat_path", ""))
        kmat_master_entry = ttk.Entry(frame, textvariable=self.kmat_master_path_var, width=55)
        kmat_master_entry.grid(row=9, column=0, sticky=(tk.W, tk.E), padx=(0, 8))

        ttk.Button(frame, text="Browse...", command=self._browse_kmat_master).grid(
            row=9, column=1, sticky=tk.W)

        # --- KMAT Export folder ---
        ttk.Label(frame, text="KMAT Export folder:", font=("Arial", 10, "bold")).grid(
            row=10, column=0, sticky=tk.W, pady=(16, 4))

        self.kmat_export_path_var = tk.StringVar(
            value=self.config.get("export_kmat_path", ""))
        kmat_export_entry = ttk.Entry(frame, textvariable=self.kmat_export_path_var, width=55)
        kmat_export_entry.grid(row=11, column=0, sticky=(tk.W, tk.E), padx=(0, 8))

        ttk.Button(frame, text="Browse...", command=self._browse_kmat_export).grid(
            row=11, column=1, sticky=tk.W)

        frame.columnconfigure(0, weight=1)

        # --- Buttons ---
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=12, column=0, columnspan=2, pady=(24, 0))

        ttk.Button(btn_frame, text="Save", command=self._save, width=12).pack(
            side=tk.LEFT, padx=(0, 12))
        ttk.Button(btn_frame, text="Cancel", command=self._cancel, width=12).pack(
            side=tk.LEFT)

        self.dialog.bind("<Return>", lambda e: self._save())
        self.dialog.bind("<Escape>", lambda e: self._cancel())

    def _browse_master(self):
        path = filedialog.askopenfilename(
            title="Select SK Master Table file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.master_path_var.set(path)

    def _browse_export(self):
        path = filedialog.askdirectory(title="Select SK Export output folder")
        if path:
            self.export_path_var.set(path)

    def _browse_zert_master(self):
        path = filedialog.askopenfilename(
            title="Select Zert Master Table file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.zert_master_path_var.set(path)

    def _browse_zert_export(self):
        path = filedialog.askdirectory(title="Select Zert Export output folder")
        if path:
            self.zert_export_path_var.set(path)

    def _browse_kmat_master(self):
        path = filedialog.askopenfilename(
            title="Select KMAT Master Table file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.kmat_master_path_var.set(path)

    def _browse_kmat_export(self):
        path = filedialog.askdirectory(title="Select KMAT Export output folder")
        if path:
            self.kmat_export_path_var.set(path)

    def _save(self):
        self.config["arbeitstabelle_path"] = self.master_path_var.get().strip()
        self.config["export_path"] = self.export_path_var.get().strip()
        self.config["arbeitstabelle_zert_path"] = self.zert_master_path_var.get().strip()
        self.config["export_zert_path"] = self.zert_export_path_var.get().strip()
        self.config["arbeitstabelle_kmat_path"] = self.kmat_master_path_var.get().strip()
        self.config["export_kmat_path"] = self.kmat_export_path_var.get().strip()
        self.result = self.config
        self.dialog.destroy()

    def _cancel(self):
        self.result = None
        self.dialog.destroy()


# ---------------------------------------------------------------------------
# NotFoundDialogSK
# ---------------------------------------------------------------------------

class NotFoundDialogSK:
    """Dialog for rolls not found in the master table.
    Pre-fills Batch No., Location and all QR dimensions; user supplies Material,
    Description, Shelf Location (mandatory) and Remarks (optional).
    Always treated as a roll — no type selector.
    """

    def __init__(self, parent, charge, qr_data):
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Roll Not Found - Manual Entry")
        self.dialog.geometry("560x640")
        self.dialog.resizable(True, True)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.geometry(
            "+%d+%d" % (parent.winfo_rootx() + 60, parent.winfo_rooty() + 60)
        )

        self._qr = qr_data  # dict from parse_qr_code
        self._build_widgets(charge)
        self.dialog.wait_window()

    def _build_widgets(self, charge):
        canvas = tk.Canvas(self.dialog, borderwidth=0)
        scrollbar = ttk.Scrollbar(self.dialog, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        inner = ttk.Frame(canvas, padding="15")
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(event):
            canvas.itemconfig(win_id, width=event.width)

        inner.bind("<Configure>", _on_frame_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        # Title
        ttk.Label(
            inner,
            text=f"Roll not found in master table!\nBatch No.: {charge}",
            font=("Arial", 12, "bold"),
            foreground="red",
            justify=tk.CENTER,
        ).pack(pady=(0, 16))

        # Pre-filled / read-only info
        info_frame = ttk.LabelFrame(inner, text="Data from QR Code (read-only)", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 12))
        info_frame.columnconfigure(1, weight=1)

        pre_filled = [
            ("Batch No.:", str(self._qr.get("charge", charge))),
            ("Location (QR):", str(self._qr.get("lort", ""))),
            ("Stage 0 - Length (mm):", str(self._qr.get("lnge0", ""))),
            ("Stage 0 - Width (mm):", str(self._qr.get("brte0", ""))),
            ("Stage 1 - Length (mm):", str(self._qr.get("lnge1", ""))),
            ("Stage 1 - Width (mm):", str(self._qr.get("brte1", ""))),
            ("Stage 2 - Length (mm):", str(self._qr.get("lnge2", ""))),
            ("Stage 2 - Width (mm):", str(self._qr.get("brte2", ""))),
        ]

        for r, (lbl, val) in enumerate(pre_filled):
            ttk.Label(info_frame, text=lbl, font=("Arial", 9, "bold")).grid(
                row=r, column=0, sticky=tk.W, pady=1)
            ttk.Label(info_frame, text=val, font=("Arial", 9)).grid(
                row=r, column=1, sticky=tk.W, padx=(8, 0), pady=1)

        # Manual fields
        manual_frame = ttk.LabelFrame(inner, text="Manual Input", padding="10")
        manual_frame.pack(fill=tk.X, pady=(0, 12))

        self.material_var = tk.StringVar()
        self.kurztext_var = tk.StringVar()
        self.fach_var = tk.StringVar()
        self.brte_meas_var = tk.StringVar()
        self.remarks_var = tk.StringVar()

        fields = [
            ("Material No. *:", self.material_var, True),
            ("Description:", self.kurztext_var, False),
            ("Shelf Location *:", self.fach_var, True),
            ("Measured Width (mm) *:", self.brte_meas_var, True),
            ("Remarks:", self.remarks_var, False),
        ]

        self._entries = {}
        for i, (label, var, _required) in enumerate(fields):
            field_frame = ttk.Frame(manual_frame)
            field_frame.pack(fill=tk.X, pady=4)
            ttk.Label(field_frame, text=label, font=("Arial", 9)).pack(anchor=tk.W)
            entry = ttk.Entry(field_frame, textvariable=var, width=50, font=("Arial", 10))
            entry.pack(fill=tk.X)
            self._entries[label] = entry
            if i == 0:
                entry.focus_set()

        # Buttons
        btn_frame = ttk.Frame(inner)
        btn_frame.pack(pady=(16, 4))
        ttk.Button(btn_frame, text="Save", command=self._save, width=12).pack(
            side=tk.LEFT, padx=(0, 12))
        ttk.Button(btn_frame, text="Cancel", command=self._cancel, width=12).pack(
            side=tk.LEFT)

        self.dialog.bind("<Escape>", lambda e: self._cancel())

    def _save(self):
        material = self.material_var.get().strip()
        fach = self.fach_var.get().strip()
        brte_meas = self.brte_meas_var.get().strip()

        if not material:
            messagebox.showerror("Error", "Material is a required field.", parent=self.dialog)
            return
        if not fach:
            messagebox.showerror("Error", "Shelf Location is a required field.", parent=self.dialog)
            return
        if not brte_meas:
            messagebox.showerror("Error", "Measured Width (mm) is a required field.", parent=self.dialog)
            return

        self.result = {
            "charge": str(self._qr.get("charge", "")),
            "lort_master": "",
            "lort_qr": str(self._qr.get("lort", "")),
            "material": material,
            "kurztext": self.kurztext_var.get().strip(),
            "werk": "",
            "lnge0": self._qr.get("lnge0", 0),
            "brte0": self._qr.get("brte0", 0),
            "lnge1": self._qr.get("lnge1", 0),
            "brte1": self._qr.get("brte1", 0),
            "lnge2": self._qr.get("lnge2", 0),
            "brte2": self._qr.get("brte2", 0),
            "flache": "",
            "frei_verw": "",
            "fach": fach,
            "brte_meas": brte_meas,
            "remarks": self.remarks_var.get().strip(),
            "status": "not_found",
        }
        self.dialog.destroy()

    def _cancel(self):
        self.result = None
        self.dialog.destroy()


# ---------------------------------------------------------------------------
# NotFoundDialogZert
# ---------------------------------------------------------------------------

class NotFoundDialogZert:
    """Dialog for Zert warehouse: charge not found in master table.
    User supplies Material No. (mandatory), Quantity (mandatory), Remarks (optional).
    """

    def __init__(self, parent, charge):
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Charge Not Found - Manual Entry")
        self.dialog.geometry("480x320")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.geometry(
            "+%d+%d" % (parent.winfo_rootx() + 60, parent.winfo_rooty() + 60)
        )
        self._charge = charge
        self._build_widgets()
        self.dialog.wait_window()

    def _build_widgets(self):
        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            frame,
            text=f"Charge not found!\nCharge: {self._charge}",
            font=("Arial", 12, "bold"),
            foreground="red",
            justify=tk.CENTER,
        ).pack(pady=(0, 16))

        info_frame = ttk.LabelFrame(frame, text="QR Data (read-only)", padding="8")
        info_frame.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(info_frame, text="Charge:", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky=tk.W)
        ttk.Label(info_frame, text=self._charge, font=("Arial", 9)).grid(row=0, column=1, sticky=tk.W, padx=(8, 0))

        manual_frame = ttk.LabelFrame(frame, text="Manual Input", padding="8")
        manual_frame.pack(fill=tk.X, pady=(0, 12))

        self.material_var = tk.StringVar()
        self.menge_var = tk.StringVar()
        self.remarks_var = tk.StringVar()

        ttk.Label(manual_frame, text="Material No. *:", font=("Arial", 9)).grid(row=0, column=0, sticky=tk.W, pady=4)
        mat_entry = ttk.Entry(manual_frame, textvariable=self.material_var, width=30, font=("Arial", 10))
        mat_entry.grid(row=0, column=1, sticky=tk.W, padx=(8, 0), pady=4)
        mat_entry.focus_set()

        ttk.Label(manual_frame, text="Recorded Quantity *:", font=("Arial", 9)).grid(row=1, column=0, sticky=tk.W, pady=4)
        menge_entry = ttk.Entry(manual_frame, textvariable=self.menge_var, width=15, font=("Arial", 10))
        menge_entry.grid(row=1, column=1, sticky=tk.W, padx=(8, 0), pady=4)

        ttk.Label(manual_frame, text="Remarks:", font=("Arial", 9)).grid(row=2, column=0, sticky=tk.W, pady=4)
        ttk.Entry(manual_frame, textvariable=self.remarks_var, width=30, font=("Arial", 10)).grid(row=2, column=1, sticky=tk.W, padx=(8, 0), pady=4)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=(8, 0))
        ttk.Button(btn_frame, text="Save", command=self._save, width=12).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Button(btn_frame, text="Cancel", command=self._cancel, width=12).pack(side=tk.LEFT)

        self.dialog.bind("<Escape>", lambda e: self._cancel())

    def _save(self):
        material = self.material_var.get().strip()
        menge = self.menge_var.get().strip()
        if not material:
            messagebox.showerror("Error", "Material No. is a required field.", parent=self.dialog)
            return
        if not menge:
            messagebox.showerror("Error", "Recorded Quantity is a required field.", parent=self.dialog)
            return
        self.result = {
            "charge": self._charge,
            "material": material,
            "kurztext": "",
            "mart": "",
            "werk": "",
            "lort": "",
            "bme": "",
            "frei_verw": "",
            "laenge_mm": "",
            "breite_mm": "",
            "adv": "",
            "menge": menge,
            "remarks": self.remarks_var.get().strip(),
            "status": "not_found",
        }
        self.dialog.destroy()

    def _cancel(self):
        self.result = None
        self.dialog.destroy()


# ---------------------------------------------------------------------------
# PositionInputDialog
# ---------------------------------------------------------------------------

class PositionInputDialog:
    """Dialog to select Position for a scanned Kaufnummer."""

    def __init__(self, parent, kauf, positions):
        self.result = None
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Position")
        self.dialog.geometry("400x200")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.geometry(
            "+%d+%d" % (parent.winfo_rootx() + 80, parent.winfo_rooty() + 80)
        )
        self._kauf = kauf
        self._positions = positions
        self._build_widgets()
        self.dialog.wait_window()

    def _build_widgets(self):
        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text=f"Kauf-Nr.: {self._kauf}",
                  font=("Arial", 12, "bold")).pack(pady=(0, 12))

        ttk.Label(frame, text="Select Position:", font=("Arial", 10)).pack(anchor=tk.W)

        self.pos_var = tk.StringVar()
        if self._positions:
            self.pos_var.set(self._positions[0])

        combo = ttk.Combobox(frame, textvariable=self.pos_var,
                             values=self._positions, font=("Arial", 12), width=15,
                             state="readonly" if self._positions else "normal")
        combo.pack(pady=(4, 16), anchor=tk.W)
        combo.focus_set()

        btn_frame = ttk.Frame(frame)
        btn_frame.pack()
        ttk.Button(btn_frame, text="OK", command=self._ok, width=10).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", command=self._cancel, width=10).pack(side=tk.LEFT)

        self.dialog.bind("<Return>", lambda e: self._ok())
        self.dialog.bind("<Escape>", lambda e: self._cancel())

    def _ok(self):
        pos = self.pos_var.get().strip()
        if not pos:
            messagebox.showerror("Error", "Please select or enter a position.", parent=self.dialog)
            return
        self.result = pos
        self.dialog.destroy()

    def _cancel(self):
        self.result = None
        self.dialog.destroy()


# ---------------------------------------------------------------------------
# NotFoundDialogKMAT
# ---------------------------------------------------------------------------

class NotFoundDialogKMAT:
    """Dialog for KMAT: Kauf+POS not found in master table."""

    def __init__(self, parent, kauf, pos):
        self.result = None
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Not Found - Manual Entry")
        self.dialog.geometry("480x300")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.geometry(
            "+%d+%d" % (parent.winfo_rootx() + 60, parent.winfo_rooty() + 60)
        )
        self._kauf = kauf
        self._pos = pos
        self._build_widgets()
        self.dialog.wait_window()

    def _build_widgets(self):
        frame = ttk.Frame(self.dialog, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            frame,
            text=f"Kauf+POS not found!\nKauf-Nr.: {self._kauf}  |  POS: {self._pos}",
            font=("Arial", 12, "bold"),
            foreground="red",
            justify=tk.CENTER,
        ).pack(pady=(0, 16))

        manual_frame = ttk.LabelFrame(frame, text="Manual Input", padding="8")
        manual_frame.pack(fill=tk.X, pady=(0, 12))

        self.material_var = tk.StringVar()
        self.menge_var = tk.StringVar()
        self.remarks_var = tk.StringVar()

        ttk.Label(manual_frame, text="Material No. *:", font=("Arial", 9)).grid(row=0, column=0, sticky=tk.W, pady=4)
        mat_entry = ttk.Entry(manual_frame, textvariable=self.material_var, width=30, font=("Arial", 10))
        mat_entry.grid(row=0, column=1, sticky=tk.W, padx=(8, 0), pady=4)
        mat_entry.focus_set()

        ttk.Label(manual_frame, text="Recorded Quantity *:", font=("Arial", 9)).grid(row=1, column=0, sticky=tk.W, pady=4)
        ttk.Entry(manual_frame, textvariable=self.menge_var, width=15, font=("Arial", 10)).grid(row=1, column=1, sticky=tk.W, padx=(8, 0), pady=4)

        ttk.Label(manual_frame, text="Remarks:", font=("Arial", 9)).grid(row=2, column=0, sticky=tk.W, pady=4)
        ttk.Entry(manual_frame, textvariable=self.remarks_var, width=30, font=("Arial", 10)).grid(row=2, column=1, sticky=tk.W, padx=(8, 0), pady=4)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(pady=(8, 0))
        ttk.Button(btn_frame, text="Save", command=self._save, width=12).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Button(btn_frame, text="Cancel", command=self._cancel, width=12).pack(side=tk.LEFT)

        self.dialog.bind("<Escape>", lambda e: self._cancel())

    def _save(self):
        material = self.material_var.get().strip()
        menge = self.menge_var.get().strip()
        if not material:
            messagebox.showerror("Error", "Material No. is required.", parent=self.dialog)
            return
        if not menge:
            messagebox.showerror("Error", "Recorded Quantity is required.", parent=self.dialog)
            return
        self.result = {
            "kauf": self._kauf,
            "pos": self._pos,
            "material": material,
            "kurztext": "",
            "werk": "",
            "lort": "",
            "bme": "",
            "frei_verw": "",
            "menge": menge,
            "remarks": self.remarks_var.get().strip(),
            "status": "not_found",
            "_mode": "kmat",
        }
        self.dialog.destroy()

    def _cancel(self):
        self.result = None
        self.dialog.destroy()


# ---------------------------------------------------------------------------
# Main Application Class
# ---------------------------------------------------------------------------

class InventurAppSK:
    """Inventory application for Forbo SK (Malacky) — Rolls only. Also supports Zert warehouse."""

    # ------------------------------------------------------------------
    # Initialisation
    # ------------------------------------------------------------------

    def __init__(self):
        self.root = tk.Tk()

        # Show warehouse selection before anything else
        self.root.withdraw()
        sel_dlg = WarehouseSelectionDialog(self.root)
        if sel_dlg.result is None:
            self.root.destroy()
            sys.exit(0)
        self.warehouse_mode = sel_dlg.result
        self.root.deiconify()

        # Base paths (EXE-compatible)
        self.base_dir = Path(self.get_base_path())
        self.config_dir = self.base_dir / "config"
        self.config_dir.mkdir(exist_ok=True)

        # Logging (needs config_dir first)
        self.setup_logging()

        # Config (may override paths)
        self.load_config()

        # Resolve arbeitstabelle / export paths from config
        self._resolve_paths()

        # Data
        self.init_data()

        # UI
        self.setup_ui()

        # Check paths & load master table
        self.check_paths_on_startup()

        # Load existing session
        if self.warehouse_mode == "Zert":
            self.load_existing_zert()
        elif self.warehouse_mode == "KMAT":
            self.load_existing_kmat()
        else:
            self.load_existing_cz()

        # Shortcuts
        self.bind_shortcuts()

    # ------------------------------------------------------------------
    # Path / Base helpers
    # ------------------------------------------------------------------

    def get_base_path(self):
        """Return base path — works for both .py script and PyInstaller .exe."""
        if getattr(sys, "frozen", False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__))

    def _resolve_paths(self):
        """Derive runtime paths from config."""
        arb = self.config.get("arbeitstabelle_path", "")
        self.arbeitstabelle_path = Path(arb) if arb else None

        exp = self.config.get("export_path", "")
        if exp:
            self.export_path = Path(exp)
        else:
            self.export_path = self.base_dir

        self.inventur_path = self.export_path / "Inventory_Rolls_SK.xlsx"
        backup_dir = self.export_path / "backups"
        try:
            backup_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        # Zert paths
        zert_arb = self.config.get("arbeitstabelle_zert_path", "")
        self.arbeitstabelle_zert_path = Path(zert_arb) if zert_arb else None

        zert_exp = self.config.get("export_zert_path", "")
        self.export_zert_path = Path(zert_exp) if zert_exp else self.base_dir
        self.inventur_zert_path = self.export_zert_path / "Inventory_Zert.xlsx"

        # KMAT paths
        kmat_arb = self.config.get("arbeitstabelle_kmat_path", "")
        self.arbeitstabelle_kmat_path = Path(kmat_arb) if kmat_arb else None

        kmat_exp = self.config.get("export_kmat_path", "")
        self.export_kmat_path = Path(kmat_exp) if kmat_exp else self.base_dir
        self.inventur_kmat_path = self.export_kmat_path / "Inventory_KMAT.xlsx"

    # ------------------------------------------------------------------
    # Logging
    # ------------------------------------------------------------------

    def setup_logging(self):
        """Configure logging to config/inventory_sk.log."""
        log_file = self.config_dir / "inventory_sk.log"
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[
                logging.FileHandler(log_file, encoding="utf-8"),
                logging.StreamHandler(),
            ],
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info(f"Inventory SK application started (mode: {getattr(self, 'warehouse_mode', 'unknown')})")

    # ------------------------------------------------------------------
    # Config
    # ------------------------------------------------------------------

    def load_config(self):
        """Load config from config/settings_sk.json."""
        config_path = self.config_dir / "settings_sk.json"
        default_config = {
            "auto_save": True,
            "arbeitstabelle_path": "",
            "export_path": "",
            "vollbild": True,
            "arbeitstabelle_zert_path": "",
            "export_zert_path": "",
            "arbeitstabelle_kmat_path": "",
            "export_kmat_path": "",
        }
        try:
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8-sig") as f:
                    self.config = json.load(f)
                # Merge any missing keys
                for k, v in default_config.items():
                    self.config.setdefault(k, v)
            else:
                self.config = default_config
                self.save_config()
        except Exception as e:
            self.config = default_config
            self.logger.error(f"Error loading config: {e}")

    def save_config(self):
        """Save config to config/settings_sk.json."""
        try:
            config_path = self.config_dir / "settings_sk.json"
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.logger.error(f"Error saving config: {e}")

    # ------------------------------------------------------------------
    # Data initialisation
    # ------------------------------------------------------------------

    def init_data(self):
        """Initialise data structures."""
        self.df_rollen = None
        self.inventur_data = []       # found rolls (SK)
        self.not_found_data = []      # not-found rolls (SK)
        self.current_scan = None
        self.current_qr_data = None
        self.undo_stack = []

        # Zert data
        self.df_zert = None
        self.inventur_data_zert = []
        self.not_found_data_zert = []

        # KMAT data
        self.df_kmat = None
        self.inventur_data_kmat = []
        self.not_found_data_kmat = []

    # ------------------------------------------------------------------
    # QR parsing
    # ------------------------------------------------------------------

    def parse_qr_code(self, raw):
        """Parse semicolon-delimited QR code.

        Format: Lort;Charge;Lnge0;Brte0;Lnge1;Brte1;Lnge2;Brte2
        Fallback: treat whole string as Charge.

        Returns dict with keys:
            lort, charge, lnge0, brte0, lnge1, brte1, lnge2, brte2
        """
        raw = raw.strip()
        if ";" in raw:
            parts = raw.split(";")

            def _int(val, default=0):
                try:
                    return int(val.strip())
                except (ValueError, AttributeError):
                    return default

            return {
                "lort": parts[0].strip() if len(parts) > 0 else "",
                "charge": parts[1].strip() if len(parts) > 1 else raw,
                "lnge0": _int(parts[2]) if len(parts) > 2 else 0,
                "brte0": _int(parts[3]) if len(parts) > 3 else 0,
                "lnge1": _int(parts[4]) if len(parts) > 4 else 0,
                "brte1": _int(parts[5]) if len(parts) > 5 else 0,
                "lnge2": _int(parts[6]) if len(parts) > 6 else 0,
                "brte2": _int(parts[7]) if len(parts) > 7 else 0,
            }
        else:
            return {
                "lort": "",
                "charge": raw,
                "lnge0": 0,
                "brte0": 0,
                "lnge1": 0,
                "brte1": 0,
                "lnge2": 0,
                "brte2": 0,
            }

    # ------------------------------------------------------------------
    # Master table (SK)
    # ------------------------------------------------------------------

    def load_arbeitstabelle(self):
        """Load SK master table from the configured path (single sheet)."""
        if self.arbeitstabelle_path is None or not self.arbeitstabelle_path.exists():
            self.df_rollen = None
            self.logger.warning("Master table not loaded (path not set or file missing)")
            return

        try:
            self.df_rollen = pd.read_excel(
                self.arbeitstabelle_path,
                dtype={"Charge": str},
            )
            if "Charge" in self.df_rollen.columns:
                self.df_rollen["Charge"] = self.df_rollen["Charge"].astype(str)

            count = len(self.df_rollen)
            self.logger.info(f"Master table loaded: {count} rows from {self.arbeitstabelle_path}")
            self.status_var.set(f"Master table loaded: {count} rolls")

            # Refresh header info label
            self._update_header_info()

        except Exception as e:
            messagebox.showerror("Error", f"Error loading master table:\n{e}")
            self.logger.error(f"Error loading master table: {e}")
            self.df_rollen = None

    # ------------------------------------------------------------------
    # Master table (Zert)
    # ------------------------------------------------------------------

    def load_arbeitstabelle_zert(self):
        """Load Zert master table from the configured path."""
        if self.arbeitstabelle_zert_path is None or not self.arbeitstabelle_zert_path.exists():
            self.df_zert = None
            self.logger.warning("Zert master table not loaded")
            return
        try:
            self.df_zert = pd.read_excel(self.arbeitstabelle_zert_path, dtype={"Charge": str})
            if "Charge" in self.df_zert.columns:
                self.df_zert["Charge"] = self.df_zert["Charge"].astype(str)
            count = len(self.df_zert)
            self.logger.info(f"Zert master table loaded: {count} rows")
            self.status_var.set(f"Zert master table loaded: {count} entries")
            self._update_header_info()
        except Exception as e:
            messagebox.showerror("Error", f"Error loading Zert master table:\n{e}")
            self.df_zert = None

    # ------------------------------------------------------------------
    # Master table (KMAT)
    # ------------------------------------------------------------------

    def load_arbeitstabelle_kmat(self):
        """Load KMAT master table from the configured path."""
        if self.arbeitstabelle_kmat_path is None or not self.arbeitstabelle_kmat_path.exists():
            self.df_kmat = None
            self.logger.warning("KMAT master table not loaded")
            return
        try:
            self.df_kmat = pd.read_excel(self.arbeitstabelle_kmat_path, dtype=str)

            def _norm_num(x):
                if x is None:
                    return ""
                s = str(x).strip()
                if s in ("", "nan"):
                    return ""
                try:
                    return str(int(float(s)))
                except (ValueError, TypeError):
                    return ""

            if "Kauf" in self.df_kmat.columns:
                self.df_kmat["Kauf"] = self.df_kmat["Kauf"].apply(_norm_num)
            if "POS" in self.df_kmat.columns:
                self.df_kmat["POS"] = self.df_kmat["POS"].apply(_norm_num)

            count = len(self.df_kmat)
            self.logger.info(f"KMAT master table loaded: {count} rows")
            self.status_var.set(f"KMAT master table loaded: {count} entries")
            self._update_header_info()
        except Exception as e:
            messagebox.showerror("Error", f"Error loading KMAT master table:\n{e}")
            self.df_kmat = None

    # ------------------------------------------------------------------
    # Charge lookup (SK)
    # ------------------------------------------------------------------

    def suche_charge(self, charge):
        """Look up a charge in df_rollen. Returns row dict or None."""
        if self.df_rollen is None:
            return None
        matches = self.df_rollen[self.df_rollen["Charge"] == str(charge)]
        if not matches.empty:
            return matches.iloc[0].to_dict()
        return None

    # ------------------------------------------------------------------
    # Charge lookup (Zert)
    # ------------------------------------------------------------------

    def suche_charge_zert(self, charge):
        """Look up a charge in df_zert. Returns row dict or None."""
        if self.df_zert is None:
            return None
        matches = self.df_zert[self.df_zert["Charge"] == str(charge)]
        if not matches.empty:
            return matches.iloc[0].to_dict()
        return None

    # ------------------------------------------------------------------
    # Kauf+POS lookup (KMAT)
    # ------------------------------------------------------------------

    def suche_kmat(self, kauf, pos):
        """Look up by Kauf + POS in df_kmat. Returns row dict or None."""
        if self.df_kmat is None:
            return None
        # Normalize input
        try:
            kauf_norm = str(int(float(str(kauf).strip())))
        except (ValueError, TypeError):
            kauf_norm = str(kauf).strip()
        try:
            pos_norm = str(int(float(str(pos).strip())))
        except (ValueError, TypeError):
            pos_norm = str(pos).strip()
        matches = self.df_kmat[
            (self.df_kmat["Kauf"] == kauf_norm) &
            (self.df_kmat["POS"] == pos_norm)
        ]
        if not matches.empty:
            return matches.iloc[0].to_dict()
        return None

    def get_kmat_positions(self, kauf):
        """Return list of available POS strings for a given Kaufnummer."""
        if self.df_kmat is None:
            return []
        try:
            kauf_norm = str(int(float(str(kauf).strip())))
        except (ValueError, TypeError):
            kauf_norm = str(kauf).strip()
        rows = self.df_kmat[self.df_kmat["Kauf"] == kauf_norm]
        return rows["POS"].tolist()

    # ------------------------------------------------------------------
    # Startup path check
    # ------------------------------------------------------------------

    def check_paths_on_startup(self):
        """Show a warning / prompt to locate file if master path is not set."""
        if self.warehouse_mode == "Zert":
            arb_missing = (
                self.arbeitstabelle_zert_path is None
                or not self.arbeitstabelle_zert_path.exists()
            )

            if arb_missing:
                answer = messagebox.askyesno(
                    "Zert Master Table Not Found",
                    "The Zert master table has not been configured or could not be found.\n\n"
                    "Search now?",
                )
                if answer:
                    path = filedialog.askopenfilename(
                        title="Select Zert Master Table",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    )
                    if path:
                        self.config["arbeitstabelle_zert_path"] = path
                        self.save_config()
                        self._resolve_paths()
                        self.load_arbeitstabelle_zert()
                    else:
                        self._disable_scan("No Zert master table selected. Scanning disabled.")
                else:
                    self._disable_scan("Zert master table not configured. Scanning disabled.")
            else:
                self.load_arbeitstabelle_zert()

            # Default export zert path
            if not self.config.get("export_zert_path", ""):
                self.config["export_zert_path"] = str(self.base_dir)
                self.save_config()
                self._resolve_paths()
        elif self.warehouse_mode == "KMAT":
            arb_missing = (
                self.arbeitstabelle_kmat_path is None
                or not self.arbeitstabelle_kmat_path.exists()
            )
            if arb_missing:
                answer = messagebox.askyesno(
                    "KMAT Master Table Not Found",
                    "The KMAT master table has not been configured or could not be found.\n\nSearch now?",
                )
                if answer:
                    path = filedialog.askopenfilename(
                        title="Select KMAT Master Table",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    )
                    if path:
                        self.config["arbeitstabelle_kmat_path"] = path
                        self.save_config()
                        self._resolve_paths()
                        self.load_arbeitstabelle_kmat()
                    else:
                        self._disable_scan("No KMAT master table selected. Scanning disabled.")
                else:
                    self._disable_scan("KMAT master table not configured. Scanning disabled.")
            else:
                self.load_arbeitstabelle_kmat()

            if not self.config.get("export_kmat_path", ""):
                self.config["export_kmat_path"] = str(self.base_dir)
                self.save_config()
                self._resolve_paths()
        else:
            arb_missing = (
                self.arbeitstabelle_path is None
                or not self.arbeitstabelle_path.exists()
            )

            if arb_missing:
                answer = messagebox.askyesno(
                    "Master Table Not Found",
                    "The master table file has not been configured or could not be found.\n\n"
                    "Would you like to locate it now?",
                )
                if answer:
                    path = filedialog.askopenfilename(
                        title="Select Master Table file",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    )
                    if path:
                        self.config["arbeitstabelle_path"] = path
                        self.save_config()
                        self._resolve_paths()
                        self.load_arbeitstabelle()
                    else:
                        self._disable_scan("No master table selected. Scanning disabled.")
                else:
                    self._disable_scan("Master table not configured. Scanning disabled.")
            else:
                self.load_arbeitstabelle()

            # Default export path to base_dir if not set
            if not self.config.get("export_path", ""):
                self.config["export_path"] = str(self.base_dir)
                self.save_config()
                self._resolve_paths()

    def _disable_scan(self, msg):
        self.scan_entry.config(state="disabled")
        self.status_var.set(f"WARNING: {msg}")
        self.logger.warning(msg)

    # ------------------------------------------------------------------
    # Settings dialog
    # ------------------------------------------------------------------

    def open_settings_dialog(self):
        """Open the settings dialog and apply any changes."""
        dlg = SettingsDialog(self.root, self.config)
        if dlg.result:
            self.config.update(dlg.result)
            self.save_config()
            self._resolve_paths()
            # Re-enable scan if it was disabled
            self.scan_entry.config(state="normal")
            if self.warehouse_mode == "Zert":
                self.load_arbeitstabelle_zert()
                messagebox.showinfo("Settings Saved", "Settings have been saved.\nZert master table reloaded.")
            elif self.warehouse_mode == "KMAT":
                self.load_arbeitstabelle_kmat()
                messagebox.showinfo("Settings Saved", "Settings have been saved.\nKMAT master table reloaded.")
            else:
                self.load_arbeitstabelle()
                messagebox.showinfo("Settings Saved", "Settings have been saved.\nMaster table reloaded.")

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def setup_ui(self):
        """Build the main UI."""
        if self.warehouse_mode == "Zert":
            self.root.title("INVENTORY Forbo - Zert Warehouse")
        elif self.warehouse_mode == "KMAT":
            self.root.title("INVENTORY Forbo - KMAT Warehouse")
        else:
            self.root.title("INVENTORY Forbo SK - Malacky Warehouse Management")
        self.root.geometry("1280x820")
        self.root.configure(bg="#f0f0f0")

        try:
            icon_path = Path(__file__).parent / "icon.ico"
            self.root.iconbitmap(str(icon_path))
        except Exception:
            pass

        self.root.state("zoomed")

        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(0, weight=1)

        self.create_header()
        self.create_scan_section()
        self.create_current_scan_section()
        self.create_list_section()
        self.create_button_section()
        self.create_status_bar()

        self.scan_entry.focus_set()

    # --- Header ---

    def create_header(self):
        header_frame = ttk.Frame(self.main_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 16))
        header_frame.columnconfigure(1, weight=1)

        if self.warehouse_mode == "Zert":
            fg_color = "#375623"
            header_text = "INVENTORY Forbo - Zert Warehouse"
        elif self.warehouse_mode == "KMAT":
            fg_color = "#6b1f1f"
            header_text = "INVENTORY Forbo - KMAT Warehouse"
        else:
            fg_color = "#1f4e79"
            header_text = "INVENTORY Forbo SK - Malacky Warehouse Management"

        ttk.Label(header_frame, text="[F]", font=("Arial", 20, "bold"),
                  foreground=fg_color).grid(row=0, column=0, padx=(0, 16))

        ttk.Label(
            header_frame,
            text=header_text,
            font=("Arial", 18, "bold"),
            foreground=fg_color,
        ).grid(row=0, column=1, sticky=tk.W)

        self.header_info_var = tk.StringVar(value="No master table loaded")
        ttk.Label(header_frame, textvariable=self.header_info_var,
                  font=("Arial", 10)).grid(row=0, column=2, padx=(20, 0))

    def _update_header_info(self):
        if self.warehouse_mode == "Zert":
            if self.df_zert is not None:
                self.header_info_var.set(f"DB: {len(self.df_zert)} entries in Zert master table")
            else:
                self.header_info_var.set("No Zert master table loaded")
        elif self.warehouse_mode == "KMAT":
            if self.df_kmat is not None:
                self.header_info_var.set(f"DB: {len(self.df_kmat)} entries in KMAT master table")
            else:
                self.header_info_var.set("No KMAT master table loaded")
        else:
            if self.df_rollen is not None:
                self.header_info_var.set(f"DB: {len(self.df_rollen)} rolls in master table")
            else:
                self.header_info_var.set("No master table loaded")

    # --- Scan section ---

    def create_scan_section(self):
        scan_frame = ttk.LabelFrame(self.main_frame, text="BARCODE / QR SCAN", padding="10")
        scan_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        scan_frame.columnconfigure(0, weight=1)

        self.scan_var = tk.StringVar()
        self.scan_entry = ttk.Entry(scan_frame, textvariable=self.scan_var,
                                    font=("Arial", 14), width=50)
        self.scan_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))

        ttk.Button(scan_frame, text="Scan", command=self.process_scan).grid(row=0, column=1)

        self.scan_entry.bind("<Return>", lambda e: self.process_scan())

        ttk.Label(
            scan_frame,
            text="Scan field is always focused. Scan a QR code or type + ENTER.",
            font=("Arial", 9),
            foreground="gray",
        ).grid(row=1, column=0, columnspan=2, pady=(5, 0))

    # --- Current scan section ---

    def create_current_scan_section(self):
        self.current_frame = ttk.LabelFrame(
            self.main_frame, text="CURRENT SCAN", padding="10")
        self.current_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        self.current_frame.columnconfigure(1, weight=1)
        self.current_frame.columnconfigure(3, weight=1)

        self._create_display_labels()
        self._create_input_panel()
        self._create_zert_info_panel()
        self._create_zert_input_panel()
        self._create_kmat_info_panel()
        self._create_kmat_input_panel()

        self.current_frame.grid_remove()

    def _create_display_labels(self):
        """Create read-only display labels for SK scan data."""
        # Row 0: Batch No. | Material
        ttk.Label(self.current_frame, text="Batch No.:", font=("Arial", 11, "bold")).grid(
            row=0, column=0, sticky=tk.W, pady=2)
        self.lbl_charge = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_charge.grid(row=0, column=1, sticky=tk.W, padx=(8, 20), pady=2)

        ttk.Label(self.current_frame, text="Material:", font=("Arial", 11, "bold")).grid(
            row=0, column=2, sticky=tk.W, pady=2)
        self.lbl_material = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_material.grid(row=0, column=3, sticky=tk.W, padx=(8, 0), pady=2)

        # Row 1: Description
        ttk.Label(self.current_frame, text="Description:", font=("Arial", 11, "bold")).grid(
            row=1, column=0, sticky=tk.W, pady=2)
        self.lbl_kurztext = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_kurztext.grid(row=1, column=1, columnspan=3, sticky=tk.W, padx=(8, 0), pady=2)

        # Row 2: Plant | Location (Master)
        ttk.Label(self.current_frame, text="Plant:", font=("Arial", 11, "bold")).grid(
            row=2, column=0, sticky=tk.W, pady=2)
        self.lbl_werk = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_werk.grid(row=2, column=1, sticky=tk.W, padx=(8, 20), pady=2)

        ttk.Label(self.current_frame, text="Location (Master):", font=("Arial", 11, "bold")).grid(
            row=2, column=2, sticky=tk.W, pady=2)
        self.lbl_lort_master = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_lort_master.grid(row=2, column=3, sticky=tk.W, padx=(8, 0), pady=2)

        # Row 3: Location QR | Area
        ttk.Label(self.current_frame, text="Location (QR):", font=("Arial", 11, "bold")).grid(
            row=3, column=0, sticky=tk.W, pady=2)
        self.lbl_lort_qr = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_lort_qr.grid(row=3, column=1, sticky=tk.W, padx=(8, 20), pady=2)

        ttk.Label(self.current_frame, text="Area (m2):", font=("Arial", 11, "bold")).grid(
            row=3, column=2, sticky=tk.W, pady=2)
        self.lbl_flache = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_flache.grid(row=3, column=3, sticky=tk.W, padx=(8, 0), pady=2)

        # Row 4: Free Usable
        ttk.Label(self.current_frame, text="Free Usable:", font=("Arial", 11, "bold")).grid(
            row=4, column=0, sticky=tk.W, pady=2)
        self.lbl_frei = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.lbl_frei.grid(row=4, column=1, sticky=tk.W, padx=(8, 0), pady=2)

        # Row 5: Dimensions grid
        self.dim_frame = ttk.LabelFrame(self.current_frame, text="Dimensions (mm)", padding="6")
        self.dim_frame.grid(row=5, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(8, 4))
        self.dim_frame.columnconfigure(1, weight=1)
        self.dim_frame.columnconfigure(3, weight=1)
        self.dim_frame.columnconfigure(5, weight=1)

        ttk.Label(self.dim_frame, text="Stage 0", font=("Arial", 10, "bold")).grid(
            row=0, column=0, columnspan=2, sticky=tk.W, padx=(0, 16))
        ttk.Label(self.dim_frame, text="Stage 1", font=("Arial", 10, "bold")).grid(
            row=0, column=2, columnspan=2, sticky=tk.W, padx=(0, 16))
        ttk.Label(self.dim_frame, text="Stage 2", font=("Arial", 10, "bold")).grid(
            row=0, column=4, columnspan=2, sticky=tk.W)

        # Stage 0
        ttk.Label(self.dim_frame, text="Length:", font=("Arial", 9)).grid(
            row=1, column=0, sticky=tk.W, padx=(0, 4))
        self.lbl_lnge0 = ttk.Label(self.dim_frame, text="", font=("Arial", 10, "bold"), width=8)
        self.lbl_lnge0.grid(row=1, column=1, sticky=tk.W, padx=(0, 16))

        # Stage 1
        ttk.Label(self.dim_frame, text="Length:", font=("Arial", 9)).grid(
            row=1, column=2, sticky=tk.W, padx=(0, 4))
        self.lbl_lnge1 = ttk.Label(self.dim_frame, text="", font=("Arial", 10, "bold"), width=8)
        self.lbl_lnge1.grid(row=1, column=3, sticky=tk.W, padx=(0, 16))

        # Stage 2
        ttk.Label(self.dim_frame, text="Length:", font=("Arial", 9)).grid(
            row=1, column=4, sticky=tk.W, padx=(0, 4))
        self.lbl_lnge2 = ttk.Label(self.dim_frame, text="", font=("Arial", 10, "bold"), width=8)
        self.lbl_lnge2.grid(row=1, column=5, sticky=tk.W)

        ttk.Label(self.dim_frame, text="Width:", font=("Arial", 9)).grid(
            row=2, column=0, sticky=tk.W, padx=(0, 4))
        self.lbl_brte0 = ttk.Label(self.dim_frame, text="", font=("Arial", 10, "bold"), width=8)
        self.lbl_brte0.grid(row=2, column=1, sticky=tk.W, padx=(0, 16))

        ttk.Label(self.dim_frame, text="Width:", font=("Arial", 9)).grid(
            row=2, column=2, sticky=tk.W, padx=(0, 4))
        self.lbl_brte1 = ttk.Label(self.dim_frame, text="", font=("Arial", 10, "bold"), width=8)
        self.lbl_brte1.grid(row=2, column=3, sticky=tk.W, padx=(0, 16))

        ttk.Label(self.dim_frame, text="Width:", font=("Arial", 9)).grid(
            row=2, column=4, sticky=tk.W, padx=(0, 4))
        self.lbl_brte2 = ttk.Label(self.dim_frame, text="", font=("Arial", 10, "bold"), width=8)
        self.lbl_brte2.grid(row=2, column=5, sticky=tk.W)

        self.sk_separator = ttk.Separator(self.current_frame, orient="horizontal")
        self.sk_separator.grid(row=6, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=8)

    def _create_input_panel(self):
        """Create user input widgets for SK (Fach + Measured Width + Remarks)."""
        self.fach_var = tk.StringVar()
        self.brte_meas_var = tk.StringVar()
        self.remarks_var = tk.StringVar()
        self.input_widgets = {}

        self.input_container = ttk.Frame(self.current_frame)
        self.input_container.grid(row=7, column=0, columnspan=4,
                                   sticky=(tk.W, tk.E), pady=4)
        self.input_container.columnconfigure(1, weight=1)

        # Fach (mandatory) — row 0
        lbl_fach = ttk.Label(self.input_container, text="Shelf Location *:",
                              font=("Arial", 11, "bold"))
        lbl_fach.grid(row=0, column=0, sticky=tk.W, pady=5)
        entry_fach = ttk.Entry(self.input_container, textvariable=self.fach_var,
                                font=("Arial", 12), width=20)
        entry_fach.grid(row=0, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        entry_fach.bind("<Return>", lambda e: self.input_widgets["brte_meas_entry"].focus_set())
        self.input_widgets["fach_entry"] = entry_fach

        # Measured Width (mandatory) — row 1
        lbl_brte_meas = ttk.Label(self.input_container, text="Measured Width (mm) *:",
                                   font=("Arial", 11, "bold"))
        lbl_brte_meas.grid(row=1, column=0, sticky=tk.W, pady=5)
        entry_brte_meas = ttk.Entry(self.input_container, textvariable=self.brte_meas_var,
                                     font=("Arial", 12), width=10)
        entry_brte_meas.grid(row=1, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        entry_brte_meas.bind("<Return>", self.save_current_scan)
        self.input_widgets["brte_meas_entry"] = entry_brte_meas

        # Remarks (optional) — row 2
        lbl_remarks = ttk.Label(self.input_container, text="Remarks (optional):",
                                 font=("Arial", 11, "bold"))
        lbl_remarks.grid(row=2, column=0, sticky=tk.W, pady=5)
        entry_remarks = ttk.Entry(self.input_container, textvariable=self.remarks_var,
                                   font=("Arial", 12), width=50)
        entry_remarks.grid(row=2, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        entry_remarks.bind("<Return>", self.save_current_scan)
        self.input_widgets["remarks_entry"] = entry_remarks

        # Save button — row 3
        save_btn = ttk.Button(self.input_container, text="Save",
                               command=self.save_current_scan)
        save_btn.grid(row=3, column=1, sticky=tk.W, padx=(10, 0), pady=8)
        self.input_widgets["save_button"] = save_btn

    def _create_zert_info_panel(self):
        """Create info display panel for Zert mode (initially hidden)."""
        self.zert_info_container = ttk.LabelFrame(
            self.current_frame, text="Zert Charge Info", padding="8")
        # Don't grid yet — shown on demand

        self.zert_info_container.columnconfigure(1, weight=1)
        self.zert_info_container.columnconfigure(3, weight=1)

        # Labels for Zert info fields
        fields_left = [
            ("Charge:", "zert_lbl_charge"),
            ("Material No.:", "zert_lbl_material"),
            ("Description:", "zert_lbl_kurztext"),
            ("Mat. Type:", "zert_lbl_mart"),
            ("Plant:", "zert_lbl_werk"),
            ("Location:", "zert_lbl_lort"),
        ]
        fields_right = [
            ("UOM:", "zert_lbl_bme"),
            ("Free Usable:", "zert_lbl_frei"),
            ("Length (mm):", "zert_lbl_laenge"),
            ("Width (mm):", "zert_lbl_breite"),
            ("ADV:", "zert_lbl_adv"),
        ]

        for r, (lbl_text, attr) in enumerate(fields_left):
            ttk.Label(self.zert_info_container, text=lbl_text,
                      font=("Arial", 10, "bold")).grid(row=r, column=0, sticky=tk.W, pady=2, padx=(0, 6))
            lbl = ttk.Label(self.zert_info_container, text="", font=("Arial", 10))
            lbl.grid(row=r, column=1, sticky=tk.W, padx=(0, 20), pady=2)
            setattr(self, attr, lbl)

        for r, (lbl_text, attr) in enumerate(fields_right):
            ttk.Label(self.zert_info_container, text=lbl_text,
                      font=("Arial", 10, "bold")).grid(row=r, column=2, sticky=tk.W, pady=2, padx=(0, 6))
            lbl = ttk.Label(self.zert_info_container, text="", font=("Arial", 10))
            lbl.grid(row=r, column=3, sticky=tk.W, pady=2)
            setattr(self, attr, lbl)

    def _create_zert_input_panel(self):
        """Create input panel for Zert mode: only Recorded Quantity."""
        self.zert_input_container = ttk.Frame(self.current_frame)
        # Don't grid yet — shown on demand

        self.menge_var = tk.StringVar()
        self.zert_remarks_var = tk.StringVar()
        self.zert_input_widgets = {}

        self.zert_input_container.columnconfigure(1, weight=1)

        ttk.Label(self.zert_input_container, text="Recorded Quantity *:",
                  font=("Arial", 11, "bold")).grid(row=0, column=0, sticky=tk.W, pady=5)
        entry_menge = ttk.Entry(self.zert_input_container, textvariable=self.menge_var,
                                font=("Arial", 12), width=15)
        entry_menge.grid(row=0, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        entry_menge.bind("<Return>", self.save_current_scan_zert)
        self.zert_input_widgets["menge_entry"] = entry_menge

        self.zert_lbl_bme_input = ttk.Label(self.zert_input_container, text="",
                                             font=("Arial", 10), foreground="gray")
        self.zert_lbl_bme_input.grid(row=0, column=2, sticky=tk.W, padx=(8, 0), pady=5)

        ttk.Label(self.zert_input_container, text="Remarks:",
                  font=("Arial", 11, "bold")).grid(row=1, column=0, sticky=tk.W, pady=5)
        entry_zert_remarks = ttk.Entry(self.zert_input_container, textvariable=self.zert_remarks_var,
                                        font=("Arial", 12), width=40)
        entry_zert_remarks.grid(row=1, column=1, columnspan=2, sticky=tk.W, padx=(10, 0), pady=5)
        entry_zert_remarks.bind("<Return>", self.save_current_scan_zert)
        self.zert_input_widgets["remarks_entry"] = entry_zert_remarks

        save_btn = ttk.Button(self.zert_input_container, text="Save",
                               command=self.save_current_scan_zert)
        save_btn.grid(row=2, column=1, sticky=tk.W, padx=(10, 0), pady=8)
        self.zert_input_widgets["save_button"] = save_btn

    def _create_kmat_info_panel(self):
        """Create info display panel for KMAT mode (initially hidden)."""
        self.kmat_info_container = ttk.LabelFrame(
            self.current_frame, text="KMAT Info", padding="8")

        self.kmat_info_container.columnconfigure(1, weight=1)
        self.kmat_info_container.columnconfigure(3, weight=1)

        fields_left = [
            ("Kauf-Nr.:", "kmat_lbl_kauf"),
            ("POS:", "kmat_lbl_pos"),
            ("Material No.:", "kmat_lbl_material"),
            ("Description:", "kmat_lbl_kurztext"),
        ]
        fields_right = [
            ("Plant:", "kmat_lbl_werk"),
            ("Location:", "kmat_lbl_lort"),
            ("UOM:", "kmat_lbl_bme"),
            ("Free Usable:", "kmat_lbl_frei"),
        ]

        for r, (lbl_text, attr) in enumerate(fields_left):
            ttk.Label(self.kmat_info_container, text=lbl_text,
                      font=("Arial", 10, "bold")).grid(row=r, column=0, sticky=tk.W, pady=2, padx=(0, 6))
            lbl = ttk.Label(self.kmat_info_container, text="", font=("Arial", 10))
            lbl.grid(row=r, column=1, sticky=tk.W, padx=(0, 20), pady=2)
            setattr(self, attr, lbl)

        for r, (lbl_text, attr) in enumerate(fields_right):
            ttk.Label(self.kmat_info_container, text=lbl_text,
                      font=("Arial", 10, "bold")).grid(row=r, column=2, sticky=tk.W, pady=2, padx=(0, 6))
            lbl = ttk.Label(self.kmat_info_container, text="", font=("Arial", 10))
            lbl.grid(row=r, column=3, sticky=tk.W, pady=2)
            setattr(self, attr, lbl)

    def _create_kmat_input_panel(self):
        """Create input panel for KMAT mode."""
        self.kmat_input_container = ttk.Frame(self.current_frame)

        self.kmat_menge_var = tk.StringVar()
        self.kmat_remarks_var = tk.StringVar()
        self.kmat_input_widgets = {}

        self.kmat_input_container.columnconfigure(1, weight=1)

        ttk.Label(self.kmat_input_container, text="Recorded Quantity *:",
                  font=("Arial", 11, "bold")).grid(row=0, column=0, sticky=tk.W, pady=5)
        entry_menge = ttk.Entry(self.kmat_input_container, textvariable=self.kmat_menge_var,
                                font=("Arial", 12), width=15)
        entry_menge.grid(row=0, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        entry_menge.bind("<Return>", self.save_current_scan_kmat)
        self.kmat_input_widgets["menge_entry"] = entry_menge

        self.kmat_lbl_bme_input = ttk.Label(self.kmat_input_container, text="",
                                             font=("Arial", 10), foreground="gray")
        self.kmat_lbl_bme_input.grid(row=0, column=2, sticky=tk.W, padx=(8, 0), pady=5)

        ttk.Label(self.kmat_input_container, text="Remarks:",
                  font=("Arial", 11, "bold")).grid(row=1, column=0, sticky=tk.W, pady=5)
        entry_remarks = ttk.Entry(self.kmat_input_container, textvariable=self.kmat_remarks_var,
                                  font=("Arial", 12), width=40)
        entry_remarks.grid(row=1, column=1, columnspan=2, sticky=tk.W, padx=(10, 0), pady=5)
        entry_remarks.bind("<Return>", self.save_current_scan_kmat)
        self.kmat_input_widgets["remarks_entry"] = entry_remarks

        save_btn = ttk.Button(self.kmat_input_container, text="Save",
                               command=self.save_current_scan_kmat)
        save_btn.grid(row=2, column=1, sticky=tk.W, padx=(10, 0), pady=8)
        self.kmat_input_widgets["save_button"] = save_btn

    def _show_sk_widgets(self):
        """Show SK-specific widgets, hide Zert and KMAT widgets."""
        # Show SK rows 0-7
        for r in range(8):
            for widget in self.current_frame.grid_slaves(row=r):
                widget.grid()
        self.dim_frame.grid()
        self.sk_separator.grid()
        self.input_container.grid()
        # Hide Zert containers
        self.zert_info_container.grid_remove()
        self.zert_input_container.grid_remove()
        # Hide KMAT containers
        self.kmat_info_container.grid_remove()
        self.kmat_input_container.grid_remove()

    def _show_zert_widgets(self):
        """Show Zert-specific widgets, hide SK and KMAT widgets."""
        # Hide all SK-specific widgets (rows 0-7 labels + dim_frame + separator + input_container)
        for widget in self.current_frame.grid_slaves():
            widget.grid_remove()

        # Show Zert containers
        self.zert_info_container.grid(row=0, column=0, columnspan=4,
                                       sticky=(tk.W, tk.E), pady=(0, 8))
        self.zert_input_container.grid(row=1, column=0, columnspan=4,
                                        sticky=(tk.W, tk.E), pady=4)

    def _show_kmat_widgets(self):
        """Show KMAT-specific widgets, hide SK and Zert widgets."""
        for widget in self.current_frame.grid_slaves():
            widget.grid_remove()
        self.kmat_info_container.grid(row=0, column=0, columnspan=4,
                                       sticky=(tk.W, tk.E), pady=(0, 8))
        self.kmat_input_container.grid(row=1, column=0, columnspan=4,
                                        sticky=(tk.W, tk.E), pady=4)

    def _hide_all_scan_widgets(self):
        """Hide all scan detail widgets."""
        self.zert_info_container.grid_remove()
        self.zert_input_container.grid_remove()
        self.kmat_info_container.grid_remove()
        self.kmat_input_container.grid_remove()
        self.input_container.grid_remove()

    # --- List section ---

    def create_list_section(self):
        list_frame = ttk.LabelFrame(self.main_frame, text="SCANNED ITEMS", padding="10")
        list_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(1, weight=1)
        self.main_frame.rowconfigure(3, weight=1)

        count_frame = ttk.Frame(list_frame)
        count_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 8))

        self.count_label = ttk.Label(count_frame, text="0 items", font=("Arial", 12, "bold"))
        self.count_label.pack(side=tk.LEFT)

        if self.warehouse_mode == "Zert":
            columns = ("Time", "Charge", "Material No.", "Description", "Quantity", "UOM", "Status")
        elif self.warehouse_mode == "KMAT":
            columns = ("Time", "Kauf-Nr.", "POS", "Material No.", "Description", "Quantity", "UOM", "Status")
        else:
            columns = ("Time", "Batch No.", "Material", "Shelf Location", "Status")

        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=12)

        for col in columns:
            self.tree.heading(col, text=col)

        if self.warehouse_mode == "Zert":
            self.tree.column("Time", width=90, minwidth=70)
            self.tree.column("Charge", width=130, minwidth=100)
            self.tree.column("Material No.", width=110, minwidth=80)
            self.tree.column("Description", width=200, minwidth=120)
            self.tree.column("Quantity", width=80, minwidth=60)
            self.tree.column("UOM", width=60, minwidth=50)
            self.tree.column("Status", width=100, minwidth=80)
        elif self.warehouse_mode == "KMAT":
            self.tree.column("Time", width=80, minwidth=70)
            self.tree.column("Kauf-Nr.", width=110, minwidth=80)
            self.tree.column("POS", width=50, minwidth=40)
            self.tree.column("Material No.", width=100, minwidth=80)
            self.tree.column("Description", width=200, minwidth=120)
            self.tree.column("Quantity", width=80, minwidth=60)
            self.tree.column("UOM", width=60, minwidth=50)
            self.tree.column("Status", width=80, minwidth=60)
        else:
            self.tree.column("Time", width=90, minwidth=70)
            self.tree.column("Batch No.", width=130, minwidth=100)
            self.tree.column("Material", width=110, minwidth=80)
            self.tree.column("Shelf Location", width=100, minwidth=70)
            self.tree.column("Status", width=130, minwidth=100)

        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        v_scroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        v_scroll.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=v_scroll.set)

        h_scroll = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        h_scroll.grid(row=2, column=0, sticky=(tk.W, tk.E))
        self.tree.configure(xscrollcommand=h_scroll.set)

        self.tree.bind("<Button-3>", self.show_context_menu)

    # --- Button section ---

    def create_button_section(self):
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(8, 0))

        ttk.Button(btn_frame, text="Export / Backup", command=self.export_backup).grid(
            row=0, column=0, padx=(0, 10))
        ttk.Button(btn_frame, text="Settings", command=self.open_settings_dialog).grid(
            row=0, column=1, padx=(0, 10))
        ttk.Button(btn_frame, text="Fullscreen (F11)", command=self.toggle_fullscreen).grid(
            row=0, column=2, padx=(0, 10))
        ttk.Button(btn_frame, text="Quit", command=self.quit_app).grid(
            row=0, column=3)

    # --- Status bar ---

    def create_status_bar(self):
        self.status_var = tk.StringVar(value="Ready to scan...")
        status_bar = ttk.Label(
            self.root, textvariable=self.status_var,
            relief=tk.SUNKEN, anchor=tk.W, font=("Arial", 9))
        status_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))

    # ------------------------------------------------------------------
    # Shortcuts
    # ------------------------------------------------------------------

    def bind_shortcuts(self):
        self.root.bind("<Control-z>", lambda e: self.undo_last_action())
        self.root.bind("<Control-Z>", lambda e: self.undo_last_action())
        self.root.bind("<Control-s>", lambda e: self._manual_save())
        self.root.bind("<Control-S>", lambda e: self._manual_save())
        self.root.bind("<Escape>", lambda e: self._reset_scan())
        self.root.bind("<F11>", lambda e: self.toggle_fullscreen())
        self.root.bind("<FocusIn>", self.ensure_scan_focus)

    # ------------------------------------------------------------------
    # Scan processing
    # ------------------------------------------------------------------

    def process_scan(self):
        """Process a scanned QR / barcode."""
        raw = self.scan_var.get().strip()
        if not raw:
            self.status_var.set("Please scan or enter a barcode / QR code.")
            return

        qr = self.parse_qr_code(raw)
        self.current_qr_data = qr
        charge = qr["charge"]

        self.logger.info(f"Scan received: raw='{raw}' -> charge='{charge}'")

        # Duplicate check
        if self._is_already_scanned(charge):
            messagebox.showwarning(
                "Already Scanned",
                f"Batch No. '{charge}' has already been scanned!\n\n"
                "Please check the list of scanned items.",
            )
            self._reset_scan()
            return

        if self.warehouse_mode == "Zert":
            row_data = self.suche_charge_zert(charge)
            if row_data is not None:
                self.show_found_zert(row_data, charge)
            else:
                self.show_not_found_dialog_zert(charge)
        elif self.warehouse_mode == "KMAT":
            kauf = raw.strip()

            # Check if Kaufnummer exists at all
            positions = self.get_kmat_positions(kauf)
            if not positions:
                messagebox.showerror(
                    "Kundenauftrag Not Found",
                    f"Kundenauftrag '{kauf}' was not found in the master table.",
                )
                self._reset_scan()
                return

            # Ask user to select position
            pos_dlg = PositionInputDialog(self.root, kauf, positions)
            if pos_dlg.result is None:
                self._reset_scan()
                return
            pos = pos_dlg.result

            # Duplicate check
            combo_key = f"{kauf}#{pos}"
            already = False
            for item in self.inventur_data_kmat + self.not_found_data_kmat:
                if f"{item.get('kauf', '')}#{item.get('pos', '')}" == combo_key:
                    already = True
                    break
            if already:
                messagebox.showwarning(
                    "Already Scanned",
                    f"Kauf-Nr. '{kauf}' / POS '{pos}' has already been scanned!",
                )
                self._reset_scan()
                return

            row_data = self.suche_kmat(kauf, pos)
            if row_data is not None:
                self.show_found_kmat(row_data, kauf, pos)
            else:
                messagebox.showerror(
                    "Position Not Found",
                    f"Position '{pos}' for Kundenauftrag '{kauf}' was not found.",
                )
                self._reset_scan()
        else:
            row_data = self.suche_charge(charge)
            if row_data is not None:
                self.show_found_rolle(row_data, qr)
            else:
                self.show_not_found_dialog(charge, qr)

    def _is_already_scanned(self, charge):
        if self.warehouse_mode == "Zert":
            for item in self.inventur_data_zert + self.not_found_data_zert:
                if str(item.get("charge", "")) == str(charge):
                    return True
            return False
        else:
            for item in self.inventur_data + self.not_found_data:
                if str(item.get("charge", "")) == str(charge):
                    return True
            return False

    # ------------------------------------------------------------------
    # Show found roll (SK)
    # ------------------------------------------------------------------

    def show_found_rolle(self, row_data, qr):
        """Display found roll data and show the input panel (SK mode)."""
        def _fmt_int(val):
            """Format numeric value as integer string (remove .0)."""
            if val == "" or val is None:
                return ""
            try:
                return str(int(float(val)))
            except (ValueError, TypeError):
                return str(val)

        def _fmt_float2(val):
            """Format float to 2 decimal places, empty if missing."""
            if val == "" or val is None:
                return ""
            try:
                return f"{float(val):.2f}"
            except (ValueError, TypeError):
                return str(val)

        # Compute Area (m2) from Länge m × Breite mm / 1000
        laenge_m = row_data.get("Länge m", None)
        breite_mm = row_data.get("Breite mm", None)
        if laenge_m not in (None, "") and breite_mm not in (None, ""):
            try:
                flache_val = f"{float(laenge_m) * float(breite_mm) / 1000:.2f}"
            except (ValueError, TypeError):
                flache_val = ""
        else:
            flache_val = ""

        self.current_scan = {
            "charge": qr["charge"],
            "lort_master": _fmt_int(row_data.get("Lagerort", "") or ""),
            "lort_qr": qr["lort"],
            "material": _fmt_int(row_data.get("Material", "") or ""),
            "kurztext": str(row_data.get("Materialkurztext", "") or ""),
            "werk": _fmt_int(row_data.get("Werk", "") or ""),
            "lnge0": qr["lnge0"],
            "brte0": qr["brte0"],
            "lnge1": qr["lnge1"],
            "brte1": qr["brte1"],
            "lnge2": qr["lnge2"],
            "brte2": qr["brte2"],
            "flache": flache_val,
            "frei_verw": _fmt_float2(row_data.get("Frei verwendbar", "") or ""),
            "status": "found",
        }

        # Update display labels
        self.lbl_charge.config(text=self.current_scan["charge"])
        self.lbl_material.config(text=self.current_scan["material"])
        self.lbl_kurztext.config(text=self.current_scan["kurztext"])
        self.lbl_werk.config(text=self.current_scan["werk"])
        self.lbl_lort_master.config(text=self.current_scan["lort_master"])
        self.lbl_lort_qr.config(text=self.current_scan["lort_qr"])
        self.lbl_flache.config(text=str(self.current_scan["flache"]))
        self.lbl_frei.config(text=str(self.current_scan["frei_verw"]))
        self.lbl_lnge0.config(text=str(self.current_scan["lnge0"]))
        self.lbl_brte0.config(text=str(self.current_scan["brte0"]))
        self.lbl_lnge1.config(text=str(self.current_scan["lnge1"]))
        self.lbl_brte1.config(text=str(self.current_scan["brte1"]))
        self.lbl_lnge2.config(text=str(self.current_scan["lnge2"]))
        self.lbl_brte2.config(text=str(self.current_scan["brte2"]))

        self.current_frame.config(text="ROLL FOUND")
        self.current_frame.grid()
        self._show_sk_widgets()

        # Reset input fields
        self.fach_var.set("")
        self.brte_meas_var.set("")
        self.remarks_var.set("")
        self.input_widgets["fach_entry"].focus_set()

        self.scan_var.set("")
        self.status_var.set(f"Roll found: {self.current_scan['kurztext']} | Enter Shelf Location")

    # ------------------------------------------------------------------
    # Show found Zert charge
    # ------------------------------------------------------------------

    def show_found_zert(self, row_data, charge):
        """Display found Zert charge data and show Zert input panel."""
        def _s(val):
            if val is None:
                return ""
            s = str(val)
            return "" if s.lower() == "nan" else s

        self.current_scan = {
            "charge": charge,
            "material": _s(row_data.get("Material", "") or _s(row_data.get("Materialnummer", ""))),
            "kurztext": _s(row_data.get("Materialkurztext", "")),
            "mart": _s(row_data.get("MArt", "")),
            "werk": _s(row_data.get("Werk", "")),
            "lort": _s(row_data.get("Lagerort", "") or _s(row_data.get("LOrt", ""))),
            "bme": _s(row_data.get("BME", "") or _s(row_data.get("Basismengeneinheit", ""))),
            "frei_verw": _s(row_data.get("Frei verwendbar", "")),
            "laenge_mm": _s(row_data.get("Länge", "") or _s(row_data.get("Länge (mm)", "") or _s(row_data.get("Laenge mm", "")))),
            "breite_mm": _s(row_data.get("Breite", "") or _s(row_data.get("Breite (mm)", "") or _s(row_data.get("Breite mm", "")))),
            "adv": _s(row_data.get("ADV", "")),
            "status": "found",
            "_mode": "zert",
        }

        # Update Zert info labels
        self.zert_lbl_charge.config(text=self.current_scan["charge"])
        self.zert_lbl_material.config(text=self.current_scan["material"])
        self.zert_lbl_kurztext.config(text=self.current_scan["kurztext"])
        self.zert_lbl_mart.config(text=self.current_scan["mart"])
        self.zert_lbl_werk.config(text=self.current_scan["werk"])
        self.zert_lbl_lort.config(text=self.current_scan["lort"])
        self.zert_lbl_bme.config(text=self.current_scan["bme"])
        self.zert_lbl_frei.config(text=self.current_scan["frei_verw"])
        self.zert_lbl_laenge.config(text=self.current_scan["laenge_mm"])
        self.zert_lbl_breite.config(text=self.current_scan["breite_mm"])
        self.zert_lbl_adv.config(text=self.current_scan["adv"])
        # Show BME hint next to menge entry
        self.zert_lbl_bme_input.config(text=self.current_scan["bme"])

        self.current_frame.config(text="CHARGE FOUND (Zert)")
        self.current_frame.grid()
        self._show_zert_widgets()

        # Reset Zert input fields
        self.menge_var.set("")
        self.zert_remarks_var.set("")
        self.zert_input_widgets["menge_entry"].focus_set()

        self.scan_var.set("")
        self.status_var.set(
            f"Charge found: {self.current_scan['kurztext']} | Enter quantity")

    # ------------------------------------------------------------------
    # Show found KMAT item
    # ------------------------------------------------------------------

    def show_found_kmat(self, row_data, kauf, pos):
        """Display found KMAT data and show KMAT input panel."""
        def _s(val):
            if val is None:
                return ""
            s = str(val)
            return "" if s.lower() == "nan" else s

        self.current_scan = {
            "kauf": kauf,
            "pos": pos,
            "material": _s(row_data.get("Materialnummer", "")),
            "kurztext": _s(row_data.get("Materialkurztext", "")),
            "werk": _s(row_data.get("Werk", "")),
            "lort": _s(row_data.get("Lagerort", "")),
            "bme": _s(row_data.get("BME", "")),
            "frei_verw": _s(row_data.get("Frei verwendbar", "")),
            "status": "found",
            "_mode": "kmat",
        }

        self.kmat_lbl_kauf.config(text=self.current_scan["kauf"])
        self.kmat_lbl_pos.config(text=self.current_scan["pos"])
        self.kmat_lbl_material.config(text=self.current_scan["material"])
        self.kmat_lbl_kurztext.config(text=self.current_scan["kurztext"])
        self.kmat_lbl_werk.config(text=self.current_scan["werk"])
        self.kmat_lbl_lort.config(text=self.current_scan["lort"])
        self.kmat_lbl_bme.config(text=self.current_scan["bme"])
        self.kmat_lbl_frei.config(text=self.current_scan["frei_verw"])
        self.kmat_lbl_bme_input.config(text=self.current_scan["bme"])

        self.current_frame.config(text="KMAT FOUND")
        self.current_frame.grid()
        self._show_kmat_widgets()

        self.kmat_menge_var.set("")
        self.kmat_remarks_var.set("")
        self.kmat_input_widgets["menge_entry"].focus_set()

        self.scan_var.set("")
        self.status_var.set(f"KMAT found: {self.current_scan['kurztext']} | Enter quantity")

    # ------------------------------------------------------------------
    # Not found dialog (SK)
    # ------------------------------------------------------------------

    def show_not_found_dialog(self, charge, qr):
        """Show dialog for rolls not in master table (SK)."""
        dlg = NotFoundDialogSK(self.root, charge, qr)

        if dlg.result:
            data = dlg.result.copy()
            data["zeitstempel"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

            self.not_found_data.append(data)

            self.undo_stack.append(("add_not_found", data.copy()))
            if len(self.undo_stack) > 50:
                self.undo_stack.pop(0)

            if self.config.get("auto_save", True):
                self.save_cz_excel()

            self.update_list()
            total = len(self.inventur_data) + len(self.not_found_data)
            self.status_var.set(
                f"Not-found roll saved. Total: {total} "
                f"({len(self.inventur_data)} found, {len(self.not_found_data)} not found)"
            )
            self.logger.info(f"Not-found roll saved: {data['charge']}")
        else:
            self.logger.info("Not-found dialog cancelled")

        self._reset_scan()

    # ------------------------------------------------------------------
    # Not found dialog (Zert)
    # ------------------------------------------------------------------

    def show_not_found_dialog_zert(self, charge):
        """Show dialog for Zert charges not in master table."""
        dlg = NotFoundDialogZert(self.root, charge)

        if dlg.result:
            data = dlg.result.copy()
            data["zeitstempel"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

            self.not_found_data_zert.append(data)

            self.undo_stack.append(("add_not_found_zert", data.copy()))
            if len(self.undo_stack) > 50:
                self.undo_stack.pop(0)

            if self.config.get("auto_save", True):
                self.save_zert_excel()

            self.update_list()
            total = len(self.inventur_data_zert) + len(self.not_found_data_zert)
            self.status_var.set(
                f"Not-found charge saved. Total: {total} "
                f"({len(self.inventur_data_zert)} found, {len(self.not_found_data_zert)} not found)"
            )
            self.logger.info(f"Not-found Zert charge saved: {data['charge']}")
        else:
            self.logger.info("Zert not-found dialog cancelled")

        self._reset_scan()

    # ------------------------------------------------------------------
    # Not found dialog (KMAT)
    # ------------------------------------------------------------------

    def show_not_found_dialog_kmat(self, kauf, pos):
        """Show dialog for KMAT Kauf+POS not in master table."""
        dlg = NotFoundDialogKMAT(self.root, kauf, pos)

        if dlg.result:
            data = dlg.result.copy()
            data["zeitstempel"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

            self.not_found_data_kmat.append(data)

            self.undo_stack.append(("add_not_found_kmat", data.copy()))
            if len(self.undo_stack) > 50:
                self.undo_stack.pop(0)

            if self.config.get("auto_save", True):
                self.save_kmat_excel()

            self.update_list()
            total = len(self.inventur_data_kmat) + len(self.not_found_data_kmat)
            self.status_var.set(
                f"Not-found KMAT saved. Total: {total} "
                f"({len(self.inventur_data_kmat)} found, {len(self.not_found_data_kmat)} not found)"
            )
            self.logger.info(f"Not-found KMAT saved: {data['kauf']}/{data['pos']}")
        else:
            self.logger.info("KMAT not-found dialog cancelled")

        self._reset_scan()

    # ------------------------------------------------------------------
    # Save current scan (SK)
    # ------------------------------------------------------------------

    def save_current_scan(self, event=None):
        """Validate and save the current found roll (SK)."""
        if not self.current_scan:
            return

        fach = self.fach_var.get().strip()
        if not fach:
            messagebox.showwarning(
                "Required Field",
                "Shelf Location is mandatory. Please enter a value.",
            )
            self.input_widgets["fach_entry"].focus_set()
            return

        brte_meas = self.brte_meas_var.get().strip()
        if not brte_meas:
            messagebox.showwarning(
                "Required Field",
                "Measured Width (mm) is mandatory. Please enter a value.",
            )
            self.input_widgets["brte_meas_entry"].focus_set()
            return

        self.current_scan["fach"] = fach
        self.current_scan["brte_meas"] = brte_meas
        self.current_scan["remarks"] = self.remarks_var.get().strip()
        self.current_scan["zeitstempel"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        self.inventur_data.append(self.current_scan.copy())

        self.undo_stack.append(("add_found", self.current_scan.copy()))
        if len(self.undo_stack) > 50:
            self.undo_stack.pop(0)

        if self.config.get("auto_save", True):
            self.save_cz_excel()

        self.update_list()

        total = len(self.inventur_data) + len(self.not_found_data)
        self.status_var.set(
            f"Roll saved. Total: {total} "
            f"({len(self.inventur_data)} found, {len(self.not_found_data)} not found)"
        )
        self.logger.info(f"Roll saved: {self.current_scan['charge']}")

        self._reset_scan()

    # ------------------------------------------------------------------
    # Save current scan (Zert)
    # ------------------------------------------------------------------

    def save_current_scan_zert(self, event=None):
        """Validate and save the current found Zert charge."""
        if not self.current_scan:
            return

        menge = self.menge_var.get().strip()
        if not menge:
            messagebox.showwarning(
                "Required Field",
                "Recorded Quantity is a required field.",
            )
            self.zert_input_widgets["menge_entry"].focus_set()
            return

        # Validate numeric
        try:
            float(menge.replace(",", "."))
        except ValueError:
            messagebox.showwarning(
                "Invalid Input",
                "Recorded Quantity must be a number.",
            )
            self.zert_input_widgets["menge_entry"].focus_set()
            return

        self.current_scan["menge"] = menge
        self.current_scan["remarks"] = self.zert_remarks_var.get().strip()
        self.current_scan["zeitstempel"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        self.inventur_data_zert.append(self.current_scan.copy())

        self.undo_stack.append(("add_found_zert", self.current_scan.copy()))
        if len(self.undo_stack) > 50:
            self.undo_stack.pop(0)

        if self.config.get("auto_save", True):
            self.save_zert_excel()

        self.update_list()

        total = len(self.inventur_data_zert) + len(self.not_found_data_zert)
        self.status_var.set(
            f"Charge saved. Total: {total} "
            f"({len(self.inventur_data_zert)} found, {len(self.not_found_data_zert)} not found)"
        )
        self.logger.info(f"Zert charge saved: {self.current_scan['charge']}")

        self._reset_scan()

    # ------------------------------------------------------------------
    # Save current scan (KMAT)
    # ------------------------------------------------------------------

    def save_current_scan_kmat(self, event=None):
        """Validate and save the current found KMAT item."""
        if not self.current_scan or self.current_scan.get("_mode") != "kmat":
            return

        menge = self.kmat_menge_var.get().strip()
        if not menge:
            messagebox.showwarning("Required Field", "Recorded Quantity is mandatory.")
            self.kmat_input_widgets["menge_entry"].focus_set()
            return

        self.current_scan["menge"] = menge
        self.current_scan["remarks"] = self.kmat_remarks_var.get().strip()
        self.current_scan["zeitstempel"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

        self.inventur_data_kmat.append(self.current_scan.copy())

        self.undo_stack.append(("add_kmat", self.current_scan.copy()))
        if len(self.undo_stack) > 50:
            self.undo_stack.pop(0)

        if self.config.get("auto_save", True):
            self.save_kmat_excel()

        self.update_list()
        total = len(self.inventur_data_kmat) + len(self.not_found_data_kmat)
        self.status_var.set(
            f"KMAT saved. Total: {total} "
            f"({len(self.inventur_data_kmat)} found, {len(self.not_found_data_kmat)} not found)"
        )
        self.logger.info(f"KMAT scan saved: {self.current_scan['kauf']}/{self.current_scan['pos']}")

        self.current_frame.grid_remove()
        self._hide_all_scan_widgets()
        self.current_scan = None
        self._reset_scan()

    # ------------------------------------------------------------------
    # Reset scan
    # ------------------------------------------------------------------

    def _reset_scan(self):
        self.current_scan = None
        self.current_qr_data = None
        self.scan_var.set("")
        self.fach_var.set("")
        self.brte_meas_var.set("")
        self.remarks_var.set("")
        if hasattr(self, "menge_var"):
            self.menge_var.set("")
        if hasattr(self, "zert_remarks_var"):
            self.zert_remarks_var.set("")
        if hasattr(self, "kmat_menge_var"):
            self.kmat_menge_var.set("")
        if hasattr(self, "kmat_remarks_var"):
            self.kmat_remarks_var.set("")
        self.current_frame.grid_remove()
        self.scan_entry.focus_set()
        self.status_var.set("Ready to scan...")

    # ------------------------------------------------------------------
    # List update
    # ------------------------------------------------------------------

    def update_list(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if self.warehouse_mode == "Zert":
            all_items = []
            for d in self.inventur_data_zert:
                all_items.append((d, "Found"))
            for d in self.not_found_data_zert:
                all_items.append((d, "Not Found"))

            all_items.sort(key=lambda x: x[0].get("zeitstempel", ""), reverse=True)

            for d, status in all_items:
                ts = d.get("zeitstempel", "")
                time_part = ts.split(" ")[1] if " " in ts else ts
                self.tree.insert(
                    "", "end",
                    values=(
                        time_part,
                        d.get("charge", ""),
                        d.get("material", ""),
                        d.get("kurztext", ""),
                        d.get("menge", ""),
                        d.get("bme", ""),
                        status,
                    ),
                )

            total = len(self.inventur_data_zert) + len(self.not_found_data_zert)
            self.count_label.config(
                text=f"{total} entries ({len(self.inventur_data_zert)} found, "
                     f"{len(self.not_found_data_zert)} not found)"
            )
        elif self.warehouse_mode == "KMAT":
            all_items = []
            for d in self.inventur_data_kmat:
                all_items.append((d, "Found"))
            for d in self.not_found_data_kmat:
                all_items.append((d, "Not Found"))

            all_items.sort(key=lambda x: x[0].get("zeitstempel", ""), reverse=True)

            for d, status in all_items:
                ts = d.get("zeitstempel", "")
                time_part = ts.split(" ")[1] if " " in ts else ts
                self.tree.insert(
                    "", "end",
                    values=(
                        time_part,
                        d.get("kauf", ""),
                        d.get("pos", ""),
                        d.get("material", ""),
                        d.get("kurztext", ""),
                        d.get("menge", ""),
                        d.get("bme", ""),
                        status,
                    ),
                )

            total = len(self.inventur_data_kmat) + len(self.not_found_data_kmat)
            self.count_label.config(
                text=f"{total} entries ({len(self.inventur_data_kmat)} found, "
                     f"{len(self.not_found_data_kmat)} not found)"
            )
        else:
            all_items = []
            for d in self.inventur_data:
                all_items.append((d, "Found"))
            for d in self.not_found_data:
                all_items.append((d, "Not Found"))

            all_items.sort(key=lambda x: x[0].get("zeitstempel", ""), reverse=True)

            for d, status in all_items:
                ts = d.get("zeitstempel", "")
                time_part = ts.split(" ")[1] if " " in ts else ts
                item_id = self.tree.insert(
                    "", "end",
                    values=(
                        time_part,
                        d.get("charge", ""),
                        d.get("material", ""),
                        d.get("fach", ""),
                        status,
                    ),
                )
                if status == "Not Found":
                    self.tree.set(item_id, "Status", "NOT FOUND")
                else:
                    self.tree.set(item_id, "Status", "Found")

            total = len(self.inventur_data) + len(self.not_found_data)
            self.count_label.config(
                text=f"{total} rolls ({len(self.inventur_data)} found, "
                     f"{len(self.not_found_data)} not found)"
            )

    # ------------------------------------------------------------------
    # Excel save (SK)
    # ------------------------------------------------------------------

    INVENTORY_HEADERS = [
        "Timestamp",
        "Plant",              # was: Werk
        "Location (Master)",  # was: LÖrt (Master)
        "Location (QR)",      # was: LÖrt (QR)
        "Material No.",       # was: Material
        "Description",        # was: Materialkurztext
        "Batch No.",          # was: Charge
        "Length S0 (mm)",     # was: Lnge
        "Width S0 (mm)",      # was: Brte
        "Length S1 (mm)",     # was: Lnge1
        "Width S1 (mm)",      # was: Brte1
        "Length S2 (mm)",     # was: Lnge2
        "Width S2 (mm)",      # was: Brte2
        "Area (m2)",          # was: Flache
        "Free Usable",        # was: +Frei verw.
        "Shelf Location",     # was: Shelf (Fach)
        "Measured Width (mm)",  # control value entered during scan
        "Remarks",
    ]

    def _row_from_item(self, d):
        def _clean(v):
            if v is None:
                return ""
            s = str(v)
            if s.lower() == "nan":
                return ""
            return s

        return [
            _clean(d.get("zeitstempel", "")),
            _clean(d.get("werk", "")),
            _clean(d.get("lort_master", "")),
            _clean(d.get("lort_qr", "")),
            _clean(d.get("material", "")),
            _clean(d.get("kurztext", "")),
            _clean(d.get("charge", "")),
            d.get("lnge0", ""),
            d.get("brte0", ""),
            d.get("lnge1", ""),
            d.get("brte1", ""),
            d.get("lnge2", ""),
            d.get("brte2", ""),
            _clean(d.get("flache", "")),
            _clean(d.get("frei_verw", "")),
            _clean(d.get("fach", "")),
            _clean(d.get("brte_meas", "")),
            _clean(d.get("remarks", "")),
        ]

    def _format_charge_col(self, ws, data_count):
        """Format Batch No. column (col 7 = G) as text."""
        charge_col = get_column_letter(7)
        for r in range(2, data_count + 2):
            ws[f"{charge_col}{r}"].number_format = "@"

    def save_cz_excel(self):
        """Write Inventory_Rolls_SK.xlsx with Inventory and Not_Found sheets."""
        try:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Inventory sheet
            ws_inv = wb.create_sheet("Inventory")
            ws_inv.append(self.INVENTORY_HEADERS)
            for d in self.inventur_data:
                ws_inv.append(self._row_from_item(d))
            self._format_charge_col(ws_inv, len(self.inventur_data))

            # Not_Found sheet
            ws_nf = wb.create_sheet("Not_Found")
            ws_nf.append(self.INVENTORY_HEADERS)
            for d in self.not_found_data:
                ws_nf.append(self._row_from_item(d))
            self._format_charge_col(ws_nf, len(self.not_found_data))

            # Ensure export dir exists
            self.export_path.mkdir(parents=True, exist_ok=True)
            wb.save(self.inventur_path)
            self.logger.info(f"Excel saved: {self.inventur_path}")

        except Exception as e:
            messagebox.showerror("Save Error", f"Error saving Excel file:\n{e}")
            self.logger.error(f"Error saving Excel: {e}")

    # ------------------------------------------------------------------
    # Excel save (Zert)
    # ------------------------------------------------------------------

    ZERT_HEADERS = [
        "Timestamp",
        "Material No.",
        "Description",
        "Mat. Type",
        "Plant",
        "Location",
        "Charge",
        "UOM",
        "Free Usable",
        "Length (mm)",
        "Width (mm)",
        "ADV",
        "Recorded Quantity",
        "Remarks",
    ]

    def _row_from_zert_item(self, d):
        def _clean(v):
            if v is None:
                return ""
            s = str(v)
            if s.lower() == "nan":
                return ""
            return s

        return [
            _clean(d.get("zeitstempel", "")),
            _clean(d.get("material", "")),
            _clean(d.get("kurztext", "")),
            _clean(d.get("mart", "")),
            _clean(d.get("werk", "")),
            _clean(d.get("lort", "")),
            _clean(d.get("charge", "")),
            _clean(d.get("bme", "")),
            _clean(d.get("frei_verw", "")),
            _clean(d.get("laenge_mm", "")),
            _clean(d.get("breite_mm", "")),
            _clean(d.get("adv", "")),
            _clean(d.get("menge", "")),
            _clean(d.get("remarks", "")),
        ]

    def save_zert_excel(self):
        """Write Inventory_Zert.xlsx with Inventory and Not_Found sheets."""
        try:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            ws_inv = wb.create_sheet("Inventory")
            ws_inv.append(self.ZERT_HEADERS)
            for d in self.inventur_data_zert:
                ws_inv.append(self._row_from_zert_item(d))

            ws_nf = wb.create_sheet("Not_Found")
            ws_nf.append(self.ZERT_HEADERS)
            for d in self.not_found_data_zert:
                ws_nf.append(self._row_from_zert_item(d))

            self.export_zert_path.mkdir(parents=True, exist_ok=True)
            wb.save(self.inventur_zert_path)
            self.logger.info(f"Zert Excel saved: {self.inventur_zert_path}")

        except Exception as e:
            messagebox.showerror("Save Error", f"Error saving Zert Excel file:\n{e}")
            self.logger.error(f"Error saving Zert Excel: {e}")

    # ------------------------------------------------------------------
    # Excel save (KMAT)
    # ------------------------------------------------------------------

    KMAT_HEADERS = [
        "Timestamp",
        "Plant",
        "Location",
        "Material No.",
        "Description",
        "Kauf-Nr.",
        "POS",
        "UOM",
        "Free Usable",
        "Recorded Quantity",
        "Remarks",
    ]

    def _row_from_kmat_item(self, d):
        def _clean(v):
            if v is None:
                return ""
            s = str(v)
            return "" if s.lower() == "nan" else s

        return [
            _clean(d.get("zeitstempel", "")),
            _clean(d.get("werk", "")),
            _clean(d.get("lort", "")),
            _clean(d.get("material", "")),
            _clean(d.get("kurztext", "")),
            _clean(d.get("kauf", "")),
            _clean(d.get("pos", "")),
            _clean(d.get("bme", "")),
            _clean(d.get("frei_verw", "")),
            _clean(d.get("menge", "")),
            _clean(d.get("remarks", "")),
        ]

    def save_kmat_excel(self):
        """Write Inventory_KMAT.xlsx with Inventory and Not_Found sheets."""
        try:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            ws_inv = wb.create_sheet("Inventory")
            ws_inv.append(self.KMAT_HEADERS)
            for d in self.inventur_data_kmat:
                ws_inv.append(self._row_from_kmat_item(d))

            ws_nf = wb.create_sheet("Not_Found")
            ws_nf.append(self.KMAT_HEADERS)
            for d in self.not_found_data_kmat:
                ws_nf.append(self._row_from_kmat_item(d))

            self.export_kmat_path.mkdir(parents=True, exist_ok=True)
            wb.save(self.inventur_kmat_path)
            self.logger.info(f"KMAT Excel saved: {self.inventur_kmat_path}")

        except Exception as e:
            messagebox.showerror("Save Error", f"Error saving KMAT Excel file:\n{e}")
            self.logger.error(f"Error saving KMAT Excel: {e}")

    # ------------------------------------------------------------------
    # Load existing session (SK)
    # ------------------------------------------------------------------

    def load_existing_cz(self):
        """Resume a previous session by loading Inventory_Rolls_SK.xlsx."""
        if not self.inventur_path.exists():
            return

        loaded = 0
        try:
            df_inv = pd.read_excel(
                self.inventur_path, sheet_name="Inventory", dtype={"Batch No.": str})
            for _, row in df_inv.iterrows():
                self.inventur_data.append(self._row_to_dict(row, status="found"))
                loaded += 1
        except Exception as e:
            self.logger.error(f"Error loading Inventory sheet: {e}")

        try:
            df_nf = pd.read_excel(
                self.inventur_path, sheet_name="Not_Found", dtype={"Batch No.": str})
            for _, row in df_nf.iterrows():
                self.not_found_data.append(self._row_to_dict(row, status="not_found"))
                loaded += 1
        except Exception:
            pass  # Sheet may not exist yet

        if loaded:
            self.update_list()
            self.status_var.set(
                f"Session resumed: {len(self.inventur_data)} found, "
                f"{len(self.not_found_data)} not found."
            )
            self.logger.info(f"Existing session loaded: {loaded} rows")

    def _row_to_dict(self, row, status):
        def _str(v):
            if pd.isna(v):
                return ""
            s = str(v)
            return "" if s.lower() == "nan" else s

        def _int_or_empty(v):
            try:
                return int(v)
            except (ValueError, TypeError):
                return 0

        return {
            "zeitstempel": _str(row.get("Timestamp", "")),
            "werk": _str(row.get("Plant", "")),
            "lort_master": _str(row.get("Location (Master)", "")),
            "lort_qr": _str(row.get("Location (QR)", "")),
            "material": _str(row.get("Material No.", "")),
            "kurztext": _str(row.get("Description", "")),
            "charge": _str(row.get("Batch No.", "")),
            "lnge0": _int_or_empty(row.get("Length S0 (mm)", 0)),
            "brte0": _int_or_empty(row.get("Width S0 (mm)", 0)),
            "lnge1": _int_or_empty(row.get("Length S1 (mm)", 0)),
            "brte1": _int_or_empty(row.get("Width S1 (mm)", 0)),
            "lnge2": _int_or_empty(row.get("Length S2 (mm)", 0)),
            "brte2": _int_or_empty(row.get("Width S2 (mm)", 0)),
            "flache": _str(row.get("Area (m2)", "")),
            "frei_verw": _str(row.get("Free Usable", "")),
            "fach": _str(row.get("Shelf Location", "")),
            "brte_meas": _str(row.get("Measured Width (mm)", "")),
            "remarks": _str(row.get("Remarks", "")),
            "status": status,
        }

    # ------------------------------------------------------------------
    # Load existing session (Zert)
    # ------------------------------------------------------------------

    def load_existing_zert(self):
        """Resume a previous Zert session by loading Inventory_Zert.xlsx."""
        if not self.inventur_zert_path.exists():
            return

        loaded = 0
        try:
            df_inv = pd.read_excel(
                self.inventur_zert_path, sheet_name="Inventory", dtype={"Charge": str})
            for _, row in df_inv.iterrows():
                self.inventur_data_zert.append(self._row_to_zert_dict(row, status="found"))
                loaded += 1
        except Exception as e:
            self.logger.error(f"Error loading Zert Inventory sheet: {e}")

        try:
            df_nf = pd.read_excel(
                self.inventur_zert_path, sheet_name="Not_Found", dtype={"Charge": str})
            for _, row in df_nf.iterrows():
                self.not_found_data_zert.append(self._row_to_zert_dict(row, status="not_found"))
                loaded += 1
        except Exception:
            pass

        if loaded:
            self.update_list()
            self.status_var.set(
                f"Zert session resumed: {len(self.inventur_data_zert)} found, "
                f"{len(self.not_found_data_zert)} not found."
            )
            self.logger.info(f"Existing Zert session loaded: {loaded} rows")

    def _row_to_zert_dict(self, row, status):
        def _str(v):
            try:
                if pd.isna(v):
                    return ""
            except (TypeError, ValueError):
                pass
            s = str(v)
            return "" if s.lower() == "nan" else s

        return {
            "zeitstempel": _str(row.get("Timestamp", "")),
            "material": _str(row.get("Material No.", "")),
            "kurztext": _str(row.get("Description", "")),
            "mart": _str(row.get("Mat. Type", "")),
            "werk": _str(row.get("Plant", "")),
            "lort": _str(row.get("Location", "")),
            "charge": _str(row.get("Charge", "")),
            "bme": _str(row.get("UOM", "")),
            "frei_verw": _str(row.get("Free Usable", "")),
            "laenge_mm": _str(row.get("Length (mm)", "")),
            "breite_mm": _str(row.get("Width (mm)", "")),
            "adv": _str(row.get("ADV", "")),
            "menge": _str(row.get("Recorded Quantity", "")),
            "remarks": _str(row.get("Remarks", "")),
            "status": status,
        }

    # ------------------------------------------------------------------
    # Load existing session (KMAT)
    # ------------------------------------------------------------------

    def load_existing_kmat(self):
        """Resume a previous KMAT session by loading Inventory_KMAT.xlsx."""
        if not self.inventur_kmat_path.exists():
            return

        loaded = 0
        try:
            df_inv = pd.read_excel(
                self.inventur_kmat_path, sheet_name="Inventory", dtype={"Kauf-Nr.": str, "POS": str})
            for _, row in df_inv.iterrows():
                self.inventur_data_kmat.append(self._row_to_kmat_dict(row, status="found"))
                loaded += 1
        except Exception as e:
            self.logger.error(f"Error loading KMAT Inventory sheet: {e}")

        try:
            df_nf = pd.read_excel(
                self.inventur_kmat_path, sheet_name="Not_Found", dtype={"Kauf-Nr.": str, "POS": str})
            for _, row in df_nf.iterrows():
                self.not_found_data_kmat.append(self._row_to_kmat_dict(row, status="not_found"))
                loaded += 1
        except Exception:
            pass

        if loaded:
            self.update_list()
            self.status_var.set(
                f"KMAT session resumed: {len(self.inventur_data_kmat)} found, "
                f"{len(self.not_found_data_kmat)} not found."
            )
            self.logger.info(f"Existing KMAT session loaded: {loaded} rows")

    def _row_to_kmat_dict(self, row, status):
        def _str(v):
            try:
                if pd.isna(v):
                    return ""
            except (TypeError, ValueError):
                pass
            s = str(v)
            return "" if s.lower() == "nan" else s

        return {
            "zeitstempel": _str(row.get("Timestamp", "")),
            "werk": _str(row.get("Plant", "")),
            "lort": _str(row.get("Location", "")),
            "material": _str(row.get("Material No.", "")),
            "kurztext": _str(row.get("Description", "")),
            "kauf": _str(row.get("Kauf-Nr.", "")),
            "pos": _str(row.get("POS", "")),
            "bme": _str(row.get("UOM", "")),
            "frei_verw": _str(row.get("Free Usable", "")),
            "menge": _str(row.get("Recorded Quantity", "")),
            "remarks": _str(row.get("Remarks", "")),
            "status": status,
            "_mode": "kmat",
        }

    # ------------------------------------------------------------------
    # Export / Backup
    # ------------------------------------------------------------------

    def export_backup(self):
        """Save a timestamped backup of the current Excel file."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            if self.warehouse_mode == "Zert":
                self.save_zert_excel()
                backup_dir = self.export_zert_path / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                backup_file = backup_dir / f"Inventory_Zert_Backup_{timestamp}.xlsx"
                source_file = self.inventur_zert_path
                if source_file.exists():
                    shutil.copy2(source_file, backup_file)
                    messagebox.showinfo(
                        "Backup Created",
                        f"Backup saved to:\n{backup_file}",
                    )
                    self.logger.info(f"Zert backup created: {backup_file}")
                else:
                    messagebox.showwarning("Warning", "No Zert data file to backup yet.")
            elif self.warehouse_mode == "KMAT":
                self.save_kmat_excel()
                backup_dir = self.export_kmat_path / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                backup_file = backup_dir / f"Inventory_KMAT_Backup_{timestamp}.xlsx"
                if self.inventur_kmat_path.exists():
                    shutil.copy2(self.inventur_kmat_path, backup_file)
                    messagebox.showinfo(
                        "Backup Created",
                        f"Backup saved to:\n{backup_file}",
                    )
                    self.logger.info(f"KMAT backup created: {backup_file}")
                else:
                    messagebox.showwarning("Warning", "No KMAT data file to backup yet.")
            else:
                self.save_cz_excel()
                backup_dir = self.export_path / "backups"
                backup_dir.mkdir(parents=True, exist_ok=True)
                backup_file = backup_dir / f"Inventory_Rolls_SK_Backup_{timestamp}.xlsx"
                if self.inventur_path.exists():
                    shutil.copy2(self.inventur_path, backup_file)
                    messagebox.showinfo(
                        "Backup Created",
                        f"Backup saved to:\n{backup_file}",
                    )
                    self.logger.info(f"Backup created: {backup_file}")
                else:
                    messagebox.showwarning("Warning", "No data file to backup yet.")
        except Exception as e:
            messagebox.showerror("Export Error", f"Error creating backup:\n{e}")
            self.logger.error(f"Backup error: {e}")

    # ------------------------------------------------------------------
    # Undo
    # ------------------------------------------------------------------

    def undo_last_action(self):
        if not self.undo_stack:
            self.status_var.set("Nothing to undo")
            return

        action, data = self.undo_stack.pop()
        charge = data.get("charge", "")

        if action == "add_found":
            self.inventur_data = [
                d for d in self.inventur_data if d.get("charge") != charge]
            if self.config.get("auto_save", True):
                self.save_cz_excel()
        elif action == "add_not_found":
            self.not_found_data = [
                d for d in self.not_found_data if d.get("charge") != charge]
            if self.config.get("auto_save", True):
                self.save_cz_excel()
        elif action == "add_found_zert":
            self.inventur_data_zert = [
                d for d in self.inventur_data_zert if d.get("charge") != charge]
            if self.config.get("auto_save", True):
                self.save_zert_excel()
        elif action == "add_not_found_zert":
            self.not_found_data_zert = [
                d for d in self.not_found_data_zert if d.get("charge") != charge]
            if self.config.get("auto_save", True):
                self.save_zert_excel()
        elif action == "add_kmat":
            kauf = data.get("kauf", "")
            pos = data.get("pos", "")
            self.inventur_data_kmat = [
                d for d in self.inventur_data_kmat
                if not (d.get("kauf") == kauf and d.get("pos") == pos)]
            if self.config.get("auto_save", True):
                self.save_kmat_excel()
        elif action == "add_not_found_kmat":
            kauf = data.get("kauf", "")
            pos = data.get("pos", "")
            self.not_found_data_kmat = [
                d for d in self.not_found_data_kmat
                if not (d.get("kauf") == kauf and d.get("pos") == pos)]
            if self.config.get("auto_save", True):
                self.save_kmat_excel()

        self.update_list()
        self.status_var.set(f"Undo: removed entry for batch '{charge}'")
        self.logger.info(f"Undo {action}: {charge}")

    # ------------------------------------------------------------------
    # Context menu
    # ------------------------------------------------------------------

    def show_context_menu(self, event):
        selection = self.tree.selection()
        if not selection:
            return
        item_id = selection[0]

        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Delete entry",
                         command=lambda: self._delete_entry(item_id))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _delete_entry(self, item_id):
        if not messagebox.askyesno("Confirm Delete",
                                   "Delete this entry permanently?"):
            return

        values = self.tree.item(item_id, "values")
        if len(values) >= 2:
            if self.warehouse_mode == "KMAT":
                # For KMAT: values = (Time, Kauf-Nr., POS, ...)
                kauf = values[1] if len(values) > 1 else ""
                pos = values[2] if len(values) > 2 else ""
                self.inventur_data_kmat = [
                    d for d in self.inventur_data_kmat
                    if not (d.get("kauf") == kauf and d.get("pos") == pos)]
                self.not_found_data_kmat = [
                    d for d in self.not_found_data_kmat
                    if not (d.get("kauf") == kauf and d.get("pos") == pos)]
                self.save_kmat_excel()
                self.update_list()
                self.status_var.set(f"Entry deleted: {kauf}/{pos}")
                self.logger.info(f"KMAT entry deleted: {kauf}/{pos}")
            else:
                charge = values[1]
                if self.warehouse_mode == "Zert":
                    self.inventur_data_zert = [
                        d for d in self.inventur_data_zert if d.get("charge") != charge]
                    self.not_found_data_zert = [
                        d for d in self.not_found_data_zert if d.get("charge") != charge]
                    self.save_zert_excel()
                else:
                    self.inventur_data = [
                        d for d in self.inventur_data if d.get("charge") != charge]
                    self.not_found_data = [
                        d for d in self.not_found_data if d.get("charge") != charge]
                    self.save_cz_excel()

                self.update_list()
                self.status_var.set(f"Entry deleted: {charge}")
                self.logger.info(f"Entry deleted: {charge}")

    # ------------------------------------------------------------------
    # Misc helpers
    # ------------------------------------------------------------------

    def toggle_fullscreen(self):
        current = self.root.attributes("-fullscreen")
        self.root.attributes("-fullscreen", not current)
        if not current:
            self.status_var.set("Fullscreen ON — press F11 to exit")
        else:
            self.status_var.set("Fullscreen OFF")

    def ensure_scan_focus(self, event=None):
        if hasattr(self, "scan_entry") and not self.current_scan:
            self.root.after(100, lambda: self.scan_entry.focus_set())

    def _manual_save(self):
        if self.warehouse_mode == "Zert":
            self.save_zert_excel()
        elif self.warehouse_mode == "KMAT":
            self.save_kmat_excel()
        else:
            self.save_cz_excel()
        self.status_var.set("Manually saved.")

    def quit_app(self):
        if messagebox.askyesno("Quit", "Are you sure you want to quit?"):
            self.logger.info("Application closed by user")
            self.root.quit()

    # ------------------------------------------------------------------
    # Run
    # ------------------------------------------------------------------

    def run(self):
        self.root.mainloop()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    try:
        app = InventurAppSK()
        app.run()
    except Exception as e:
        messagebox.showerror("Critical Error", f"Unexpected error:\n{e}")
        logging.error(f"Critical error: {e}")
        raise


if __name__ == "__main__":
    main()
